"""
Содержит родительские классы экономических отчетов
"""
from decimal import Decimal
import xml.dom.minidom
import os
import shutil
import zipfile
import openpyxl
import pytils
import sys

class Settings:
    """
    Настройки отчетов
    """
    
    def __init__(self, path_to_settings_file):
        """
        Инициализация класса
        """
        
        # self.path_to_file = path_to_settings_file
        self.__load_settings(path_to_settings_file)
        
        # Инициализируем словарь данных страховых компаний
        self.__var_insurance_company_data()
        # Инициализируем словарь данных медицинской организации
        self.__var_medical_organization_data()
        # Загружаем данные страховых компаний
        self.__load_insurance_company_data(self.__path_to_dir_insurance_company_data)
        # Путь к папке с данными медицинской организации
        # Загружаем данные медицинской организации
        self.__load_medical_organization_data(self.__path_to_dir_medical_organization_data)
        self.__load_profil_ambulance(self.path_profiles_ambulance)
        self.__load_profil_hospital(self.path_profiles_hospital)
    
    def __load_settings(self, path_to_file_setting):
        """
        загружаем настройки
        """
        try:
            self.source_file = xml.dom.minidom.parse(path_to_file_setting)
            self.source_file.normalize()
        except FileNotFoundError:
            return 'Файл '+ path_to_file_setting + ' не существует'
        
        # Папка для временных файлов, должна отличаться от source_dir
        self.work_dir = self.__parse('work_dir')
        # Полный путь к файлу с разбивкой по страховым
        self.path_to_work_file = ''
        # Сетевая папка из которой берутся файлы реестра и выкладываются результаты
        self.source_dir = self.__parse('source_dir')
        # Полный код МО по тфомс
        self.code_lpu_in_tfoms = self.__parse('code_lpu_in_tfoms')
        # Первые 5 цифр кода лпу по ТФОМС для поиска файлов реестров
        self.code_lpu_for_seach_files = self.code_lpu_in_tfoms[:5]
        # Полное наименование МО
        self.full_name_lpu = self.__parse('full_name_lpu')
        # Признак круглосуточного стационара
        self.priznak_ks = int(self.__parse('priznak_ks'))
        # Признак дневного стационара
        self.priznak_ds = int(self.__parse('priznak_ds'))
        # Путь к каталогу с файлами настроек
        self.settings_file_path = self.__parse('settings_file_path')
        # Файл содержащий данные для формирования счета в страховые
        # общая для всех форма счета (таблица и подписи)
        self.bill_format_xls = self.__parse('bill_format_xls')
        # Файл с форматом отчета сводного счета поликлиники по страховой компании
        self.consolidated_ambulance_insurance_company_format_xls = self.__parse('consolidated_ambulance_insurance_company_format_xls')
        # Файл с форматом отчета сводного счета стационара по страховой компании
        self.consolidated_hospital_insurance_company_format_xls = self.__parse('consolidated_hospital_insurance_company_format_xls')
        # Первый этап диспансеризации взрослых
        self.dispanser_adult_I = self.__parse('dispanser_adult_I')
        # Второй этап диспансеризации взрослых
        self.dispanser_adult_II = self.__parse('dispanser_adult_II')
        # Первый этап диспансеризации детей-сирот в стационаре
        self.dispanser_children_stacionar_I = self.__parse('dispanser_children_stacionar_I')
        # Второй этап диспансеризации детей-сирот в стационаре
        self.dispanser_children_stacionar_II = self.__parse('dispanser_children_stacionar_II')
        # Первый этап диспансеризации детей-сирот
        self.dispanser_children_I = self.__parse('dispanser_children_I')
        # Второй этап диспансеризации детей-сирот
        self.dispanser_children_II = self.__parse('dispanser_children_II')
        # Профилактические медицинские осмотры взростого населения
        self.profosmotr_adult = self.__parse('profosmotr_adult')
        # Первый этап профилактических медицинских осмотров несовершеннолетних
        self.profosmotr_children_I = self.__parse('profosmotr_children_I')
        # Второй этап профилактических медицинских осмотров несовершеннолетних
        self.profosmotr_children_II = self.__parse('profosmotr_children_II')
        # Профиль стоматологической поликлиники
        self.profil_stomatolog = self.__parse('profil_stomatolog')
        # Признак готовности к подготовке счетов и отчетов после нахождения
        # файла с разбивкой по страховым принимает значение True
        self.bill_go = False
        # Признак отладки
        if self.__parse('debug') == 'True':
            self.debug = True
        else:
            self.debug = False
        # Путь к файлу с профилями полклинических отделений
        self.path_profiles_ambulance = self.__parse('profiles_ambulance')
        # Путь к файлу с профилями стацонарных отделений
        self.path_profiles_hospital = self.__parse('profiles_hospital')
        # Путь к папке с данными страховых компаний
        self.__path_to_dir_insurance_company_data = self.__parse('path_to_dir_insurance_company_data')
        self.__path_to_dir_medical_organization_data = self.__parse('path_to_dir_medical_organization_data')
        self.code_tfoms = self.__parse('code_tfoms')
    
    def __var_insurance_company_data(self):
        """
        Инициализация словаря даных страховых компаний для формирования счетов
        """
        self.insurance_company_data={'kod_smo': [],
                                     'full_name_smo': [],
                                     'address_smo': [],
                                     'address_smo_ext': [],
                                     'telephone_smo': [],
                                     'personal_account_smo': [],
                                     'checking_account_smo': [],
                                     'single_treasury_account_smo': [],
                                     'treasury_account_smo': [],
                                     'payer_bank': [],
                                     'bik_smo': [],
                                     'inn_smo': [],
                                     'kpp_smo': [],
                                     'okotmo_smo': []}
    
    def __var_medical_organization_data(self):
        """
        Инициализация словаря данных медицинской организации для формирования счетов
        """
        self.medical_organization_data={'kod_mo': None,
                                        'full_name_mo': None,
                                        'address_mo': None,
                                        'address_mo_ext': None,
                                        'telephone_mo': None,
                                        'personal_account_mo': None,
                                        'checking_account_mo': None,
                                        'checking_account_mo_tfoms': None,
                                        'correspondent_account_mo': None,
                                        'single_treasury_account_mo': None,
                                        'treasury_account_mo': None,
                                        'payee_bank': None,
                                        'payee_bank_for_tfoms': None,
                                        'kbk_mo': None,
                                        'bik_mo': None,
                                        'inn_mo': None,
                                        'kpp_mo': None,
                                        'okotmo_mo': None}
    
    def __parse(self, value):
        
        try:
            search_value = self.source_file.getElementsByTagName(value)[0]
        except IndexError:
            return 'Файл не содержит тэг: ' + value
        else:
            return search_value.childNodes[0].nodeValue

    def search_work_files(self):
        for name in os.listdir(path=str(self.source_dir)):
            if name.find(self.code_lpu_for_seach_files) != -1 and name.find('.zip') != -1:
                self.path_to_work_file = self.source_dir + name
                self.__clear_work_dir()
                self.__unzip_work_file(self.path_to_work_file, self.work_dir)
                self.__unzip_smo_files()
                if not self.debug:
                    os.remove(self.path_to_work_file)
                self.bill_go = True
                return self.path_to_work_file
        return False
    
    def __search_data_xml_files(self, path_to_dir):
        """
        Поиск файлов с данными страховых компаний и медиционской организации
        Возвращает список найденных файлов
        """
        path_to_files = []
        for name in os.listdir(path=str(path_to_dir)):
            if name.find('.xml') != -1:
                path_to_files.append(path_to_dir+'/'+name)
        path_to_files.sort()
        return path_to_files
    
    def __load_insurance_company_data(self, path_to_dir_insurance_company_data):
        """
        Загрузка данных страховых компаний для формирвания счета
        """
        for path_to_file_insurance_company_data in self.__search_data_xml_files(path_to_dir_insurance_company_data):
            file_insurance_company_data = xml.dom.minidom.parse(path_to_file_insurance_company_data)
            try:
                insurance_company_data = file_insurance_company_data.getElementsByTagName("insurance_company_data")
            except IndexError:
                continue
            # Зургужаем код страховой компании
            try:
                self.insurance_company_data['kod_smo'].append(insurance_company_data[0].getElementsByTagName("kod_smo")[0].childNodes[0].data)
            except IndexError:
                self.insurance_company_data['kod_smo'].append('')
            # Загружаем полное наименование страховой компании
            try:
                self.insurance_company_data['full_name_smo'].append(insurance_company_data[0].getElementsByTagName("full_name_smo")[0].childNodes[0].data)
            except IndexError:
                self.insurance_company_data['full_name_smo'].append('')
            # Загружаем адрес страховой компании
            try:
                self.insurance_company_data['address_smo'].append(insurance_company_data[0].getElementsByTagName("address_smo")[0].childNodes[0].data)
            except IndexError:
                self.insurance_company_data['address_smo'].append('')
            # Загружаем доп информаци об адресе страховой компании
            try:
                self.insurance_company_data['address_smo_ext'].append(insurance_company_data[0].getElementsByTagName("address_smo_ext")[0].childNodes[0].data)
            except IndexError:
                self.insurance_company_data['address_smo_ext'].append('')
            # Загружаем номера телефонов страховой компании
            try:
                self.insurance_company_data['telephone_smo'].append(insurance_company_data[0].getElementsByTagName("telephone_smo")[0].childNodes[0].data)
            except IndexError:
                self.insurance_company_data['telephone_smo'].append('')
            # Загружаем лицевой счет страховой компании
            try:
                self.insurance_company_data['personal_account_smo'].append(insurance_company_data[0].getElementsByTagName("personal_account_smo")[0].childNodes[0].data)
            except IndexError:
                self.insurance_company_data['personal_account_smo'].append('')
            # Загружаем расчетный счет страховой компании
            try:
                self.insurance_company_data['checking_account_smo'].append(insurance_company_data[0].getElementsByTagName("checking_account_smo")[0].childNodes[0].data)
            except IndexError:
                self.insurance_company_data['checking_account_smo'].append('')
            # Загружаем единый казначейский счет страховой компании
            try:
                self.insurance_company_data['single_treasury_account_smo'].append(insurance_company_data[0].getElementsByTagName("single_treasury_account_smo")[0].childNodes[0].data)
            except IndexError:
                self.insurance_company_data['single_treasury_account_smo'].append('')
            # Загружаем казначейский счет страховой компании
            try:
                self.insurance_company_data['treasury_account_smo'].append(insurance_company_data[0].getElementsByTagName("treasury_account_smo")[0].childNodes[0].data)
            except IndexError:
                self.insurance_company_data['treasury_account_smo'].append('')
            # Загружаем онформацию о банке страховой компании
            try:
                self.insurance_company_data['payer_bank'].append(insurance_company_data[0].getElementsByTagName("payer_bank")[0].childNodes[0].data)
            except IndexError:
                self.insurance_company_data['payer_bank'].append('')
            # Загружаем бик страховой компании
            try:
                self.insurance_company_data['bik_smo'].append(insurance_company_data[0].getElementsByTagName("bik_smo")[0].childNodes[0].data)
            except IndexError:
                self.insurance_company_data['bik_smo'].append('')
            # Загружаем ИНН страховой компании
            try:
                self.insurance_company_data['inn_smo'].append(insurance_company_data[0].getElementsByTagName("inn_smo")[0].childNodes[0].data)
            except IndexError:
                self.insurance_company_data['inn_smo'].append('')
            # Загружаем КПП страховой компании
            try:
                self.insurance_company_data['kpp_smo'].append(insurance_company_data[0].getElementsByTagName("kpp_smo")[0].childNodes[0].data)
            except IndexError:
                self.insurance_company_data['kpp_smo'].append('')
            # Загружаем ОКОТМО страховой сомпании
            try:
                self.insurance_company_data['okotmo_smo'].append(insurance_company_data[0].getElementsByTagName("okotmo_smo")[0].childNodes[0].data)
            except IndexError:
                self.insurance_company_data['okotmo_smo'].append('')

    def __load_medical_organization_data(self, path_to_dir_medical_organization_data):
        """
        Загрузка данных медицинской организации
        """
        for path_to_file_medical_organization_data in self.__search_data_xml_files(path_to_dir_medical_organization_data):
            file_medical_organization_data = xml.dom.minidom.parse(path_to_file_medical_organization_data)
            try:
                medical_organization_data = file_medical_organization_data.getElementsByTagName("medical_organization_data")
            except IndexError:
                continue
            # Зургужаем код Медицинской организации
            try:
                self.medical_organization_data['kod_mo'] = medical_organization_data[0].getElementsByTagName("kod_mo")[0].childNodes[0].data
            except IndexError:
                self.medical_organization_data['kod_mo'] = ''
            # Загружаем полное наименование Медицинской организации
            try:
                self.medical_organization_data['full_name_mo'] = medical_organization_data[0].getElementsByTagName("full_name_mo")[0].childNodes[0].data
            except IndexError:
                self.medical_organization_data['full_name_mo'] = ''
            # Загружаем адрес Медицинской организации
            try:
                self.medical_organization_data['address_mo'] = medical_organization_data[0].getElementsByTagName("address_mo")[0].childNodes[0].data
            except IndexError:
                self.medical_organization_data['address_mo'] = ''
            # Загружаем доп информаци об адресе Медицинской организации
            try:
                self.medical_organization_data['address_mo_ext'] = medical_organization_data[0].getElementsByTagName("address_mo_ext")[0].childNodes[0].data
            except IndexError:
                self.medical_organization_data['address_mo_ext'] = ''
            # Загружаем номера телефонов Медицинской организации
            try:
                self.medical_organization_data['telephone_mo'] = medical_organization_data[0].getElementsByTagName("telephone_mo")[0].childNodes[0].data
            except IndexError:
                self.medical_organization_data['telephone_mo'] = ''
            # Загружаем лицевой счет Медицинской организации
            try:
                self.medical_organization_data['personal_account_mo'] = medical_organization_data[0].getElementsByTagName("personal_account_mo")[0].childNodes[0].data
            except IndexError:
                self.medical_organization_data['personal_account_mo'] = ''
            # Загружаем расчетный счет Медицинской организации
            try:
                self.medical_organization_data['checking_account_mo'] = medical_organization_data[0].getElementsByTagName("checking_account_mo")[0].childNodes[0].data
            except IndexError:
                self.medical_organization_data['checking_account_mo'] = ''
            # Загружаем расчетный счет Медицинской организации для счета ТФОМС
            try:
                self.medical_organization_data['checking_account_mo_tfoms'] = medical_organization_data[0].getElementsByTagName("checking_account_mo_tfoms")[0].childNodes[0].data
            except IndexError:
                self.medical_organization_data['checking_account_mo_tfoms'] = ''
            # Загружаем корреспондентский счет Медицинской организации
            try:
                self.medical_organization_data['correspondent_account_mo'] = medical_organization_data[0].getElementsByTagName("correspondent_account_mo")[0].childNodes[0].data
            except IndexError:
                self.medical_organization_data['correspondent_account_mo'] = ''
            # Загружаем единый казначейский счет Медицинской организации
            try:
                self.medical_organization_data['single_treasury_account_mo'] = medical_organization_data[0].getElementsByTagName("single_treasury_account_mo")[0].childNodes[0].data
            except IndexError:
                self.medical_organization_data['single_treasury_account_mo'] = ''
            # Загружаем казначейский счет Медицинской организации
            try:
                self.medical_organization_data['treasury_account_mo'] = medical_organization_data[0].getElementsByTagName("treasury_account_mo")[0].childNodes[0].data
            except IndexError:
                self.medical_organization_data['treasury_account_mo'] = ''
            # Загружаем онформацию о банке Медицинской организации
            try:
                self.medical_organization_data['payee_bank'] = medical_organization_data[0].getElementsByTagName("payee_bank")[0].childNodes[0].data
            except IndexError:
                self.medical_organization_data['payee_bank'] = ''
            # Загружаем онформацию о банке Медицинской организации для тфомс
            try:
                self.medical_organization_data['payee_bank_for_tfoms'] = medical_organization_data[0].getElementsByTagName("payee_bank_for_tfoms")[0].childNodes[0].data
            except IndexError:
                self.medical_organization_data['payee_bank_for_tfoms'] = ''
            # Загружаем КБК Медицинской организации
            try:
                self.medical_organization_data['kbk_mo'] = medical_organization_data[0].getElementsByTagName("kbk_mo")[0].childNodes[0].data
            except IndexError:
                self.medical_organization_data['kbk_mo'] = ''
            # Загружаем бик Медицинской организации
            try:
                self.medical_organization_data['bik_mo'] = medical_organization_data[0].getElementsByTagName("bik_mo")[0].childNodes[0].data
            except IndexError:
                self.medical_organization_data['bik_mo'] = ''
            # Загружаем ИНН Медицинской организации
            try:
                self.medical_organization_data['inn_mo'] = medical_organization_data[0].getElementsByTagName("inn_mo")[0].childNodes[0].data
            except IndexError:
                self.medical_organization_data['inn_mo'] = ''
            # Загружаем КПП Медицинской организации
            try:
                self.medical_organization_data['kpp_mo'] = medical_organization_data[0].getElementsByTagName("kpp_mo")[0].childNodes[0].data
            except IndexError:
                self.medical_organization_data['kpp_mo'] = ''
            # Загружаем ОКОТМО Медицинской организации
            try:
                self.medical_organization_data['okotmo_mo'] = medical_organization_data[0].getElementsByTagName("okotmo_mo")[0].childNodes[0].data
            except IndexError:
                self.medical_organization_data['okotmo_mo'] = ''
    
    def __clear_work_dir(self):
        """
        Очистка временной папаки от всего содержимого
        На входе необходимо указать путь к папке для очистки
        """
        
        for name in os.listdir (path = str (self.work_dir)):
            file_for_delete = self.work_dir + name
            try:
                os.remove(file_for_delete)
            except IsADirectoryError:
                shutil.rmtree(file_for_delete)
            except PermissionError:
                shutil.rmtree(file_for_delete)
        return True
    
    def __unzip_work_file(self, source_file, distenation_path):
        """
        Распаковка при условии наличия исходного файла
        """
        tfom_zip = zipfile.ZipFile (source_file)
        tfom_zip.extractall(distenation_path)
        tfom_zip.close
    
    def __unzip_smo_files(self):
        """
        Распаковываем архив во временную папку с файлами 
        счетов для страховых компаний
        """
        list_f = os.listdir(path=str(self.work_dir))
        for name in list_f:
            file_smo = self.work_dir + name
            if (name.find(self.code_lpu_for_seach_files) != -1 and
                name.find('.zip') != -1 and name.find('_err') == -1 and
                name.find('_99') == -1):
                self.bill_go = True
                dir_out = self.work_dir + os.path.splitext(name)[0]
                os.mkdir(dir_out)
                self.__unzip_work_file(file_smo, dir_out)
            os.remove(file_smo)

    def __load_profil_ambulance(self, path_to_file_xml):
        """
        Загружаем словарь врачебных профилей поликлиники
        """
        self.dict_ambulance_profiles = {}
        file_profiles = self.source_file = xml.dom.minidom.parse(path_to_file_xml)
        
        fields = file_profiles.getElementsByTagName("zap")
        
        for zap in fields:
            self.dict_ambulance_profiles[zap.getElementsByTagName("prof")[0].childNodes[0].data] = zap.getElementsByTagName("name")[0].childNodes[0].data

    def __load_profil_hospital(self, path_to_file_xml):
        """
        Загружаем словарь профилей стационарной помощи
        """
        self.dict_hospital_profiles = {}
        file_profiles = self.source_file = xml.dom.minidom.parse(path_to_file_xml)
        
        fields = file_profiles.getElementsByTagName("zap")
        
        for zap in fields:
            self.dict_hospital_profiles[zap.getElementsByTagName("prof")[0].childNodes[0].data] = zap.getElementsByTagName("name")[0].childNodes[0].data


class DataSMO:
    """
    Данные полученные из страховой компании используемые
    для формирования счетов и отчетов
    """
    # Название меяцев для даты счета
    _months = ['января',
              'февраля',
              'марта',
              'апреля',
              'мая',
              'июня',
              'июля',
              'августа',
              'сентября',
              'октября',
              'ноября',
              'декабря']
    _months_now = ['январь',
             'февраль',
             'март',
             'апрель',
             'май',
             'июнь',
             'июль',
             'август',
             'сентябрь',
             'октябрь',
             'ноябрь',
             'декабрь']

    
    def __init__(self, path_to_xml_file, settings):
        xml_source_file = xml.dom.minidom.parse(path_to_xml_file)
        bill_info = xml_source_file.getElementsByTagName("SCHET")
        cases = xml_source_file.getElementsByTagName("SLUCH")
        # Инициализируем переменные для формирования счета на оплату услуг
        self.__var_for_bill(bill_info)
        # Инициализируем перемернные для сводного счета поликлиники
        self.__var_consolidated_bill_ambulance()
        # Инициализируем перемернные для сводного счета круглосуточного стационара
        self.__var_consolidated_bill_ks()
        # Инициализируем перемернные для сводного счета дневного стационара
        self.__var_consolidated_bill_ds()
        # Инициализируем словарь сводного счета поликлиники
        self.__var_consolidated_ambulance_insurance_company()
        # Инициализируем словарь сводного счета круглосуточного стационара
        self.__var_consolidated_ks_insurance_company()
        # Инициализируем словарь сводного счета дневного стационара
        self.__var_consolidated_ds_insurance_company()
        # Собираем данные для формирования счета на оплату
        if self.bill_data['kod_smo'] != '61010':
            self.__data_for_bill(cases, settings)
        else:
            self.__data_for_bill_tfoms(cases, settings)
        # Сумма счета прописью
        self.__capitalize()
        # Собираем данные для сводного счета по поликлинике
        self.__consolidated_bill_ambulance(cases, settings)
        # Собираем данные для сводного счета полкилники по страховой компании
        # с разбивкой по профилям оказания и поликлиникам
        self.__consolidated_bill_ambulance_insurance_company(cases, settings)
        # Сбор данных для сводного счета по круглосуточному стационару
        self.__consolidated_bill_ks(cases, settings)
        # Сбор данных для сводного счета по дневному стационару
        self.__consolidated_bill_ds(cases, settings)
        # Сбор данных для сводного счета по профилям круглосуточного стационара
        self.__consolidated_bill_ks_insurance_company(cases, settings)
        # Сбор данных для сводного счета по профилям дневного стационара
        self.__consolidated_bill_ds_insurance_company(cases, settings)
        # Синхронизируем одинаковые переменные в разных словарях
        self.consolidated_ambulance['current_month'] = self._months_now[self.bill_data['month_bill'] - 1]
        self.consolidated_ambulance['number_consolidated'] = self.bill_data['month_bill']
        self.consolidated_ambulance['year_bill'] = self.bill_data['year_bill']
        self.consolidated_ambulance['kod_smo'] = self.bill_data['kod_smo']
        self.consolidated_ambulance_insurance_company['current_month'] = self._months_now[self.bill_data['month_bill'] - 1]
        self.consolidated_ambulance_insurance_company['number_consolidated'] = self.bill_data['month_bill']
        self.consolidated_ambulance_insurance_company['year_bill'] = self.bill_data['year_bill']
        self.consolidated_ambulance_insurance_company['kod_smo'] = self.bill_data['kod_smo']
        self.consolidated_ks_insurance_company['type_hospital'] = settings.priznak_ks
        self.consolidated_ks_insurance_company['current_month'] = self._months_now[self.bill_data['month_bill'] - 1]
        self.consolidated_ks_insurance_company['number_consolidated'] = self.bill_data['month_bill']
        self.consolidated_ks_insurance_company['year_bill'] = self.bill_data['year_bill']
        self.consolidated_ks_insurance_company['kod_smo'] = self.bill_data['kod_smo']
        self.consolidated_ds_insurance_company['type_hospital'] = settings.priznak_ds
        self.consolidated_ds_insurance_company['current_month'] = self._months_now[self.bill_data['month_bill'] - 1]
        self.consolidated_ds_insurance_company['number_consolidated'] = self.bill_data['month_bill']
        self.consolidated_ds_insurance_company['year_bill'] = self.bill_data['year_bill']
        self.consolidated_ds_insurance_company['kod_smo'] = self.bill_data['kod_smo']

    def __var_consolidated_bill_ambulance(self):
        """
        Переменные для сводного счета по поликлинике
        """
        self.consolidated_ambulance = {'kod_smo': '',                   # Код страховой компании
                                       'current_month': '',             # Месяц подготовки сводного счета
                                       'year_bill': '',                 # Отчетный год
                                       'number_consolidated': '',       # Номер сводного, он же номер месяца
                                       'visits': Decimal(0),            # Количество амбулаторных посещений для сводного счета поликлиники
                                       'appeal': Decimal(0),            # Количество амбулаторных обращений для сводного счета поликлиники
                                       'service': Decimal(0),           # Количество амбулаторных услуг для сводного счета поликлиники
                                       'dentistry_uet': Decimal(0),     # Количество УЕТ стоматологии для сводного счета поликлиники
                                       'summ': Decimal(0)}              # Сумма для сводного счета поликлиники

    def __var_consolidated_ambulance_insurance_company(self):
        """
        Инициализируем словарь сводного счета поликлиники
        """
        self.consolidated_ambulance_insurance_company = {'kod_smo': '',             # Код страховой компании
                                                         'current_month': '',       # Месяц подготовки сводного счета
                                                         'year_bill': '',           # Отчетный год
                                                         'number_consolidated': '', # Номер сводного, он же номер месяца
                                                         'kod_lpu': [],             # Код ЛПУ оказания
                                                         'profil': [],              # Профиль отделения оказания помощи
                                                         'kod_usl': [],             # Код посещения, услуги
                                                         'visits': [],              # Количество посещений
                                                         'appeal': [],              # Количество обращений
                                                         'services_in_sluch': [],   # Количество услуг в случае
                                                         'dentistry_uet': [],       # Количество УЕТ стоматологии
                                                         'individual_bill': [],     # Количество индивидуальных счетов
                                                         'summ_sluch': []}          # Стоимость услуги

    def __var_consolidated_ks_insurance_company(self):
        """
        Инициализируем словарь сводного счета круглосуточного стационара
        """
        self.consolidated_ks_insurance_company = {'kod_smo': '',             # Код страховой компании
                                                  'number_value': 1,
                                                  'current_month': '',       # Месяц подготовки сводного счета
                                                  'year_bill': '',           # Отчетный год
                                                  'number_consolidated': '', # Номер сводного, он же номер месяца
                                                  'kod_lpu': [],             # 0 Код ЛПУ оказания
                                                  'podr': [],                # 1 Код подразделения оказания помощи
                                                  'pacients': [],            # 3 Колчисевто выбывших пациентов
                                                  'amount_of_days': [],      # 4 Количество койкодней фактическое
                                                  'amount_of_days_paid': [], # 5 Количество койкодней оплачено
                                                  'fksg': [],                # 2 Код примененного КСГ, поле code_usl
                                                  'summ': [],                # 6 Сумма за оказанную помощь
                                                  'type_hospital': ''}       # 7 Вид стационарной помощий
    
    def __var_consolidated_ds_insurance_company(self):
        """
        Инициализируем словарь сводного счета дневного стационара
        """
        self.consolidated_ds_insurance_company = {'kod_smo': '',             # Код страховой компании
                                                  'number_value': 1,
                                                  'current_month': '',       # Месяц подготовки сводного счета
                                                  'year_bill': '',           # Отчетный год
                                                  'number_consolidated': '', # Номер сводного, он же номер месяца
                                                  'kod_lpu': [],             # 0 Код ЛПУ оказания
                                                  'podr': [],                # 1 Код подразделения оказания помощи
                                                  'pacients': [],            # 3 Колчисевто выбывших пациентов
                                                  'amount_of_days': [],      # 4 Количество койкодней фактическое
                                                  'amount_of_days_paid': [], # 5 Количество койкодней оплачено
                                                  'fksg': [],                # 2 Код примененного КСГ, поле code_usl
                                                  'summ': [],                # 6 Сумма за оказанную помощь
                                                  'type_hospital': ''}       # 7 Вид стационарной помощий

    def __var_for_bill(self, bill_info):
        """
        Переменные для формирования счета
        """
        
        self.bill_data = {'kod_smo': '',                                                       # Код страховой компании в ТФОМС
                          'number_bill': '',                                                    # Номер счета
                          'date_bill': '',                                                      # Дата выставления счета
                          'month_bill': '',                                                     # Месяц выставления счета
                          'year_bill': '',                                                      # Год выставления счета
                          'current_month': '',                                                  # Текущий месяц
                          'past_month': '',                                                     # Предыдущий месяц
                          'ks_current_month_summ': Decimal(0),                                  # Сумма по круглосуточному стационару в текущем месяце
                          'ks_previous_month_summ': Decimal(0),                                 # Сумма по круглосуточному стационару в предыдущем месяце
                          'ds_current_month_summ': Decimal(0),                                  # Сумма по дневному стационару в текущем месяце
                          'ds_previous_month_summ': Decimal(0),                                 # Сумма по дневному стационару в предыдущем месяце
                          'ambulance_current_month_summ': Decimal(0),                           # Сумма по амбулаторной помощи в текущем месяце
                          'ambulance_previous_month_summ': Decimal(0),                          # Сумма по амбулаторной помощи в предыдущем месяце
                          'smp_current_month_summ': Decimal(0),                                 # Сумма по скорой помощи в текущем месяце
                          'smp_thrombolysis_current_month_summ': Decimal(0),                    # Сумма по скорой помощи за вызовы с тромболизисом в текущем месяце
                          'smp_previous_month_summ': Decimal(0),                                # Сумма по скорой помощи в предыдущем месяце
                          'smp_thrombolysis_previous_month_summ': Decimal(0),                   # Сумма по скорой помощи за вызовы с тромболизисом в предыдущем месяце
                          'capitalize': '',                                                     # Сумма счета прописью
                          'summ_pf': Decimal(0),                                                # Сумма подушевого финансирования
                          'summ_smp': Decimal(0),                                               # Сумма подушевого финансирования СМП
                          'summ_fap': Decimal(0),                                               # Сумма подушевого финансирования ФАП
                          'dispanser_adult_I_current_month_summ': Decimal(0),                   # Сумма за диспансеризацию I этап взрослого населения в текущем месяце
                          'dispanser_adult_I_previous_month_summ': Decimal(0),                  # Сумма за диспансеризацию I этап взрослого населения в предыдущем месяце
                          'dispanser_adult_II_current_month_summ': Decimal(0),                  # Сумма за диспансеризацию II этап взрослого населения в текущем месяце
                          'dispanser_adult_II_previous_month_summ': Decimal(0),                 # Сумма за диспансеризацию II этап взрослого населения в предыдущем месяце
                          'profosmotr_adult_current_month_summ': Decimal(0),                    # Сумма за профосмотр взрослого населения в текущем месяце
                          'profosmotr_adult_previous_month_summ': Decimal(0),                   # Сумма за профосмотр взрослого населения в предыдущем месяце
                          'dispanser_children_stacionar_I_current_month_summ': Decimal(0),      # Сумма за диспансеризацию I этап детей сирот в трудной жизненной ситуации в текущем месяце
                          'dispanser_children_stacionar_I_previous_month_summ': Decimal(0),     # Сумма за диспансеризацию I этап детей сирот в трудной жизненной ситуации в предцдущем месяце
                          'dispanser_children_stacionar_II_current_month_summ': Decimal(0),     # Сумма за диспансеризацию II этап детей сирот в трудной жизненной ситуации в текущем месяце
                          'dispanser_children_stacionar_II_previous_month_summ': Decimal(0),    # Сумма за диспансеризацию II этап детей сирот в трудной жизненной ситуации в предыдущем месяце
                          'dispanser_children_I_current_month_summ': Decimal(0),                # Сумма за диспансеризацию I этап детей сирот без попечительства родителей в текущем месяце
                          'dispanser_children_I_previous_month_summ': Decimal(0),               # Сумма за диспансеризацию I этап детей сирот без попечительства родителей в предыдущем месяце
                          'dispanser_children_II_current_month_summ': Decimal(0),               # Сумма за диспансеризацию II этап детей сирот без попечительства родителей в текущем месяце
                          'dispanser_children_II_previous_month_summ': Decimal(0),              # Сумма за диспансеризацию II этап детей сирот без попечительства родителей в предыдущем месяце
                          'profosmotr_children_current_month_summ': Decimal(0),                 # Сумма за профосмотр детей в текущем месяце
                          'profosmotr_children_previous_month_summ': Decimal(0)                 # Сумма за профосмотр детей в предыдцщем месяце
                         }

        # Код страховой компании в ТФОМС
        self.bill_data['kod_smo'] = bill_info[0].getElementsByTagName("PLAT")[0].childNodes[0].data
        # Номер счета
        self.bill_data['number_bill'] = bill_info[0].getElementsByTagName("NSCHET")[0].childNodes[0].data
        # Дата выставления счета
        self.bill_data['date_bill'] = bill_info[0].getElementsByTagName("DSCHET")[0].childNodes[0].data
        self.bill_data['date_bill'] = self.__format_date(self.bill_data['kod_smo'], self.bill_data['date_bill'])
        # Месяц выставления счета
        self.bill_data['month_bill'] = int(bill_info[0].getElementsByTagName("MONTH")[0].childNodes[0].data)
        # Текущий месяц
        self.bill_data['current_month'] = self._months_now[self.bill_data['month_bill'] - 1]
        # Предыдущий месяц
        self.bill_data['past_month'] = self._months_now[self.bill_data['month_bill'] - 2] if self.bill_data['month_bill']-1 > 1 else self._months_now[11]
        # Год выставления счета
        self.bill_data['year_bill'] = bill_info[0].getElementsByTagName("YEAR")[0].childNodes[0].data
        
        
        if self.bill_data['kod_smo'] != '61010':
            # Сумма подушевого финансирования
            self.bill_data['summ_pf'] = Decimal(bill_info[0].getElementsByTagName("SUMMA_PF")[0].childNodes[0].data)
            # Сумма подушевого финансирования Скорой помощи
            self.bill_data['summ_smp'] = Decimal(bill_info[0].getElementsByTagName("SUMMA_SMP")[0].childNodes[0].data)
            self.bill_data['smp_current_month_summ'] += self.bill_data['summ_smp']
            # Сумма подушевого финансирования ФАП
            self.bill_data['summ_fap'] = Decimal(bill_info[0].getElementsByTagName("SUMMA_FAP")[0].childNodes[0].data)

    def __var_consolidated_bill_ks(self):
        """
        Переменные для сводного счета по круглосуточному стационару
        """
        self.consolidated_ks = {'pacients': Decimal(0),             # Колчиство пациентов, получивших помощь
                                'amount_of_days': Decimal(0),       # Количество койкодней проведенных в стационаре
                                'amount_of_days_paid': Decimal(0),  # Количество койкодней учтенных при выставлении счета
                                'summ': Decimal(0)}                 # Сумма за круглосуточный стационар

    def __format_date(self, kod_smo, date):
        """
        для всех кроме тфомс формиование даты счета в формате дд.мм.гггг
        для тфомс формиование даты счета в формате дд месяц гггг
        """
        if kod_smo != '61010':
            result = date[8] + date[9] + '.' + date[5] + date[6] + '.'
            for i in range (4):
                result += date[i]
        else:
            result = date[8] + date[9] + ' ' + self._months[int(date[5] + date[6]) - 1] + ' '
            for i in range (4):
                result += date[i]
        return result

    def __var_consolidated_bill_ds(self):
        """
        Переменные для сводного счета по дневному стационару
        """
        self.consolidated_ds = {'pacients': Decimal(0),             # Колчиство пациентов, получивших помощь
                                'amount_of_days': Decimal(0),       # Количество койкодней проведенных в стационаре
                                'amount_of_days_paid': Decimal(0),  # Количество койкодней учтенных при выставлении счета
                                'summ': Decimal(0)}                 # Сумма за круглосуточный стационар

    def ready(self):
        # Проверка на создание объекта
        return True

    def __data_for_bill(self, cases, settings):
        """
        Данные для счета в страховю компанию кроме ТФОМС
        """
        
        for case in cases:
            nsvod = int(case.getElementsByTagName("NSVOD")[0].childNodes[0].data)
            summ_case = Decimal(case.getElementsByTagName("SUMV")[0].childNodes[0].data)
            
            svod_type = nsvod//100
            
            try:
                dispancer_case = case.getElementsByTagName("DISP_SL")[0]
                dispancer = dispancer_case.childNodes[0].data
            except IndexError:
                dispancer = False
            
            if svod_type == 1:  # Круглосуточный стационар
                if nsvod-100 == int(self.bill_data['month_bill']):
                    self.bill_data['ks_current_month_summ'] += summ_case
                else:
                    self.bill_data['ks_previous_month_summ'] += summ_case
            elif svod_type == 2:    # Дневной стационар
                if nsvod-200 == int(self.bill_data['month_bill']):
                    self.bill_data['ds_current_month_summ'] += summ_case
                else:
                    self.bill_data['ds_previous_month_summ'] += summ_case
            elif svod_type == 3:    # Амбулаторная помощь
                if nsvod-300 == int(self.bill_data['month_bill']):
                    self.bill_data['ambulance_current_month_summ'] += summ_case
                    if dispancer:
                        if dispancer == settings.dispanser_adult_I:
                            self.bill_data['dispanser_adult_I_current_month_summ'] += summ_case
                        elif dispancer == settings.dispanser_adult_II:
                            self.bill_data['dispanser_adult_II_current_month_summ'] += summ_case
                        elif dispancer == settings.profosmotr_adult:
                            self.bill_data['profosmotr_adult_current_month_summ'] += summ_case
                        elif dispancer == settings.dispanser_children_stacionar_I:
                            self.bill_data['dispanser_children_stacionar_I_current_month_summ'] += summ_case
                        elif dispancer == settings.dispanser_children_stacionar_II:
                            self.bill_data['dispanser_children_stacionar_II_current_month_summ'] += summ_case
                        elif dispancer == settings.dispanser_children_I:
                            self.bill_data['dispanser_children_I_current_month_summ'] += summ_case
                        elif dispancer == settings.profosmotr_children_I or dispancer == settings.profosmotr_children_II:
                            self.bill_data['profosmotr_children_current_month_summ'] += summ_case
                else:
                    self.bill_data['ambulance_previous_month_summ'] += summ_case
                    if dispancer:
                        if dispancer == settings.dispanser_adult_I:
                            self.bill_data['dispanser_adult_I_previous_month_summ'] += summ_case
                        elif dispancer == settings.dispanser_adult_II:
                            self.bill_data['dispanser_adult_II_previous_month_summ'] += summ_case
                        elif dispancer == settings.profosmotr_adult:
                            self.bill_data['profosmotr_adult_previous_month_summ'] += summ_case
                        elif dispancer == settings.dispanser_children_stacionar_I:
                            self.bill_data['dispanser_children_stacionar_I_previous_month_summ'] += summ_case
                        elif dispancer == settings.dispanser_children_stacionar_II:
                            self.bill_data['dispanser_children_stacionar_II_previous_month_summ'] += summ_case
                        elif dispancer == settings.dispanser_children_I:
                            self.bill_data['dispanser_children_I_previous_month_summ'] += summ_case
                        elif dispancer == settings.profosmotr_children_I or dispancer == settings.profosmotr_children_II:
                            self.bill_data['profosmotr_children_previous_month_summ'] += summ_case
            else:                   # Скорая помощь
                if nsvod-400 == int(self.bill_data['month_bill']):
                    self.bill_data['smp_current_month_summ'] += summ_case
                else:
                    self.bill_data['smp_previous_month_summ'] += summ_case

    def __data_for_bill_tfoms(self, cases, settings):
        """
        Данные для счета в ТФОМС
        """
        
        for case in cases:
            nsvod = int(case.getElementsByTagName("NSVOD")[0].childNodes[0].data)
            summ_case = Decimal(case.getElementsByTagName("SUMV")[0].childNodes[0].data)
            
            svod_type = nsvod//100
            
            if svod_type == 1:  # Круглосуточный стационар
                if nsvod-100 == int(self.bill_data['month_bill']):
                    self.bill_data['ks_current_month_summ'] += summ_case
                else:
                    self.bill_data['ks_previous_month_summ'] += summ_case
            elif svod_type == 2:    # Дневной стационар
                if nsvod-200 == int(self.bill_data['month_bill']):
                    self.bill_data['ds_current_month_summ'] += summ_case
                else:
                    self.bill_data['ds_previous_month_summ'] += summ_case
            elif svod_type == 3:    # Амбулаторная помощь
                if nsvod-300 == int(self.bill_data['month_bill']):
                    self.bill_data['ambulance_current_month_summ'] += summ_case
                else:
                    self.bill_data['ambulance_previous_month_summ'] += summ_case
            else:                   # Скорая помощь
                if nsvod-400 == int(self.bill_data['month_bill']):
                    self.bill_data['smp_current_month_summ'] += summ_case
                else:
                    self.bill_data['smp_previous_month_summ'] += summ_case

    def __calculation_uet(self, usl):
        """
        Расчет УЕТ за 1 услугу стоматологии
        """
        usl_summ = float(usl.getElementsByTagName("SUMV_USL")[0].childNodes[0].data)
        usl_tarif = float(usl.getElementsByTagName("TARIF")[0].childNodes[0].data)
        uet_stomatolog = int(usl_summ/usl_tarif*100)
        return Decimal(uet_stomatolog)/100
    
    def __consolidated_bill_ambulance(self, cases, settings):
        """
        Сбор данных для сводного счета по поликлинике
        """
        
        for case in cases:
            services_in_case = 0    # Для каждого случая необходимо обнулять количество услуг
            nsvod = int(case.getElementsByTagName("NSVOD")[0].childNodes[0].data)
            summ_case = Decimal(case.getElementsByTagName("SUMV")[0].childNodes[0].data)
            if nsvod//100 == 3:
                self.consolidated_ambulance['summ'] += summ_case
                
                services_in_case = len(case.getElementsByTagName("USL"))
                
                try:
                    dispancer_case = case.getElementsByTagName("DISP_SL")[0]
                    dispancer = dispancer_case.childNodes[0].data
                except IndexError:
                    dispancer = False
                
                podr = case.getElementsByTagName("PODR")[0].childNodes[0].data
                if podr == settings.profil_stomatolog:
                    usl_stomatolog = case.getElementsByTagName("USL")
                    for usl in usl_stomatolog:
                        self.consolidated_ambulance['dentistry_uet'] += self.__calculation_uet(usl)
                elif services_in_case == 1 or dispancer:
                    self.consolidated_ambulance['visits'] += 1
                else:
                    self.consolidated_ambulance['appeal'] += 1 if services_in_case != 0 else 0
                    self.consolidated_ambulance['service'] += services_in_case

    def __consolidated_bill_ambulance_insurance_company(self, cases, settings):
        """
        Собираем данные для сводного счета полкилники по страховой компании
        с разбивкой по профилям оказания и поликлиникам
        """
        
        for case in cases:
            nsvod = int(case.getElementsByTagName("NSVOD")[0].childNodes[0].data)
            
            if nsvod//100 == 3:
                consolidated_sluch = {'kod_lpu': None,          # 0 Код ЛПУ оказания
                                      'profil': None,           # 1 Профиль отделения оказания помощи
                                      'kod_usl': None,          # 2 Код посещения, услуги
                                      'visits': 0,              # 3 Количество посещений
                                      'appeal': 0,              # 4 Количество обращений
                                      'services_in_sluch': 0,   # 5 Количество услуг в случае
                                      'dentistry_uet': 0,       # 6 Количество УЕТ стоматологии
                                      'individual_bill': 1,     # 7 Количество индивидуальных счетов
                                      'summ_sluch': 0}          # 8 Стоимость услуги
                # Стоимость случая
                consolidated_sluch['summ_sluch'] += Decimal(case.getElementsByTagName("SUMV")[0].childNodes[0].data)
                # Код ЛПУ оказания 
                consolidated_sluch['kod_lpu'] = int(case.getElementsByTagName("KODLPU")[0].childNodes[0].data)
                # Профиль отделения оказания помощи
                consolidated_sluch['profil'] = int(case.getElementsByTagName("PODR")[0].childNodes[0].data)
                for usl in case.getElementsByTagName("USL"):
                    # Количество услуг в случае
                    consolidated_sluch['services_in_sluch'] += 1
                    if consolidated_sluch['profil'] == int(settings.profil_stomatolog):
                        # Указвыаем код услуги, в итоге будет учтена только последняя
                        consolidated_sluch['kod_usl'] = usl.getElementsByTagName("CODE_USL")[0].childNodes[0].data
                        # производим расет УЕТ
                        consolidated_sluch['dentistry_uet'] += self.__calculation_uet(usl)
                        # Указываем тип визита стоматолог
                        type_visit = 0
                    elif (usl.getElementsByTagName("IDMASTER")[0].childNodes[0].data ==
                          usl.getElementsByTagName("IDSERV")[0].childNodes[0].data):
                        # Указываем код главной услуги
                        consolidated_sluch['kod_usl'] = usl.getElementsByTagName("CODE_USL")[0].childNodes[0].data
                        # Указываем тип визита 1 - посещение 2 - обращение
                        type_visit = 1 if (int(consolidated_sluch['kod_usl']) // 10**6 - consolidated_sluch['profil'] * 10 != 2) else 2
                
                if type_visit == 1:
                    consolidated_sluch['visits'] = 1
                elif type_visit == 2:
                    consolidated_sluch['appeal'] = 1
                
                self.__consolidated_ambulance_add_finded_case(consolidated_sluch)

    def __consolidated_ambulance_add_finded_case(self, consolidated_sluch):
        """
        Добавляем информацию о найденном случае в общий словарь поликлиники
        """
        address_cell_kod_lpu = -1
        address_cell_profil = -1
        address_cell_kod_usl = -1
        
        # Ищем ЛПУ в словаре
        if consolidated_sluch['kod_lpu'] in self.consolidated_ambulance_insurance_company['kod_lpu']:
            # ЛПУ есть в словаре, ищем позицию
            for cell_kod_lpu in range(len(self.consolidated_ambulance_insurance_company['kod_lpu'])):
                if self.consolidated_ambulance_insurance_company['kod_lpu'][cell_kod_lpu] == consolidated_sluch['kod_lpu']:
                    address_cell_kod_lpu = cell_kod_lpu
                    break
        else:
            # ЛПУ нет в словаре, добавляем и запоминаем позицию
            address_cell_kod_lpu = self.__consolidated_ambulance_insurance_company_add_new_lpu(consolidated_sluch['kod_lpu'], consolidated_sluch['profil'], consolidated_sluch['kod_usl'])
        
        # Ищем профиль в словаре
        if address_cell_profil == -1 and (consolidated_sluch['profil'] in self.consolidated_ambulance_insurance_company['profil'][address_cell_kod_lpu]):
            # профиль есть в списке, ищем позицию
            for cell_profil in range(0, len(self.consolidated_ambulance_insurance_company['profil'][address_cell_kod_lpu])):
                if self.consolidated_ambulance_insurance_company['profil'][address_cell_kod_lpu][cell_profil] == consolidated_sluch['profil']:
                    address_cell_profil = cell_profil
                    break
        else:
            # профиля нет в списке, добавляем
            address_cell_profil = self.__consolidated_ambulance_insurance_company_add_new_profil(consolidated_sluch['profil'], consolidated_sluch['kod_usl'], address_cell_kod_lpu)
        
        # Ищем код услуги в списке
        if address_cell_kod_usl == -1 and consolidated_sluch['kod_usl'] in self.consolidated_ambulance_insurance_company['kod_usl'][address_cell_kod_lpu][address_cell_profil]:
            # код услуги есть в списке, ищем позицию
            for cell_kod_usl in range(len(self.consolidated_ambulance_insurance_company['kod_usl'][address_cell_kod_lpu][address_cell_profil])):
                if self.consolidated_ambulance_insurance_company['kod_usl'][address_cell_kod_lpu][address_cell_profil][cell_kod_usl] == consolidated_sluch['kod_usl']:
                    address_cell_kod_usl = cell_kod_usl
                    break
        else:
            # кода услуги нет в списке, добавляем
            address_cell_kod_usl = self.__consolidated_ambulance_insurance_company_add_new_usl(consolidated_sluch['kod_usl'], address_cell_kod_lpu, address_cell_profil)
        # прибавляем данные для лп, профиля, кода услуги
        self.consolidated_ambulance_insurance_company['visits'][address_cell_kod_lpu][address_cell_profil][address_cell_kod_usl] += consolidated_sluch['visits']
        self.consolidated_ambulance_insurance_company['appeal'][address_cell_kod_lpu][address_cell_profil][address_cell_kod_usl] += consolidated_sluch['appeal']
        self.consolidated_ambulance_insurance_company['services_in_sluch'][address_cell_kod_lpu][address_cell_profil][address_cell_kod_usl] += consolidated_sluch['services_in_sluch']
        self.consolidated_ambulance_insurance_company['dentistry_uet'][address_cell_kod_lpu][address_cell_profil][address_cell_kod_usl] += consolidated_sluch['dentistry_uet']
        self.consolidated_ambulance_insurance_company['individual_bill'][address_cell_kod_lpu][address_cell_profil][address_cell_kod_usl] += consolidated_sluch['individual_bill']
        self.consolidated_ambulance_insurance_company['summ_sluch'][address_cell_kod_lpu][address_cell_profil][address_cell_kod_usl] += consolidated_sluch['summ_sluch']

    def __consolidated_ambulance_insurance_company_add_new_lpu(self, kod_lpu, profil, kod_usl):
        """
        Добавляем новую больницу в общий словарь поликлиники
        """
        self.consolidated_ambulance_insurance_company['kod_lpu'].append(kod_lpu)
        address_cell_kod_lpu = len(self.consolidated_ambulance_insurance_company['kod_lpu']) - 1
        self.consolidated_ambulance_insurance_company['profil'].append([])
        self.consolidated_ambulance_insurance_company['kod_usl'].append([])
        self.consolidated_ambulance_insurance_company['visits'].append([])
        self.consolidated_ambulance_insurance_company['appeal'].append([])
        self.consolidated_ambulance_insurance_company['services_in_sluch'].append([])
        self.consolidated_ambulance_insurance_company['dentistry_uet'].append([])
        self.consolidated_ambulance_insurance_company['individual_bill'].append([])
        self.consolidated_ambulance_insurance_company['summ_sluch'].append([])
        self.__consolidated_ambulance_insurance_company_add_new_profil(profil, kod_usl, address_cell_kod_lpu)
        return address_cell_kod_lpu
    
    def __consolidated_ambulance_insurance_company_add_new_profil(self, profil, kod_usl, address_cell_kod_lpu):
        """
        Добавляем новое подразделение в общий словарь поликлиники
        """
        self.consolidated_ambulance_insurance_company['profil'][address_cell_kod_lpu].append(profil)
        address_cell_profil = len(self.consolidated_ambulance_insurance_company['profil'][address_cell_kod_lpu]) - 1
        self.consolidated_ambulance_insurance_company['kod_usl'][address_cell_kod_lpu].append([])
        self.consolidated_ambulance_insurance_company['visits'][address_cell_kod_lpu].append([])
        self.consolidated_ambulance_insurance_company['appeal'][address_cell_kod_lpu].append([])
        self.consolidated_ambulance_insurance_company['services_in_sluch'][address_cell_kod_lpu].append([])
        self.consolidated_ambulance_insurance_company['dentistry_uet'][address_cell_kod_lpu].append([])
        self.consolidated_ambulance_insurance_company['individual_bill'][address_cell_kod_lpu].append([])
        self.consolidated_ambulance_insurance_company['summ_sluch'][address_cell_kod_lpu].append([])
        self.__consolidated_ambulance_insurance_company_add_new_usl(kod_usl, address_cell_kod_lpu, address_cell_profil)
        return address_cell_profil
    
    def __consolidated_ambulance_insurance_company_add_new_usl(self, kod_usl, address_cell_kod_lpu, address_cell_profil):
        """
        Добавляем новую услугу в подразделение в общий словарь поликлиники
        """
        self.consolidated_ambulance_insurance_company['kod_usl'][address_cell_kod_lpu][address_cell_profil].append(kod_usl)
        self.consolidated_ambulance_insurance_company['visits'][address_cell_kod_lpu][address_cell_profil].append(0)
        self.consolidated_ambulance_insurance_company['appeal'][address_cell_kod_lpu][address_cell_profil].append(0)
        self.consolidated_ambulance_insurance_company['services_in_sluch'][address_cell_kod_lpu][address_cell_profil].append(0)
        self.consolidated_ambulance_insurance_company['dentistry_uet'][address_cell_kod_lpu][address_cell_profil].append(0)
        self.consolidated_ambulance_insurance_company['individual_bill'][address_cell_kod_lpu][address_cell_profil].append(0)
        self.consolidated_ambulance_insurance_company['summ_sluch'][address_cell_kod_lpu][address_cell_profil].append(0)
        return len(self.consolidated_ambulance_insurance_company['kod_usl'][address_cell_kod_lpu][address_cell_profil]) - 1

    def __consolidated_bill_ks(self, cases, settings):
        """
        Сбор данных для сводного счета по круглосуточному стационару
        """
        
        for case in cases:
            nsvod = int(case.getElementsByTagName("NSVOD")[0].childNodes[0].data)
            if nsvod//100 == settings.priznak_ks:
                self.consolidated_ks['amount_of_days'] += int(case.getElementsByTagName("KD_Z")[0].childNodes[0].data)
                for usl in case.getElementsByTagName("USL"):
                    try:
                        self.consolidated_ks['amount_of_days_paid'] += int(usl.getElementsByTagName("KD")[0].childNodes[0].data)
                    except IndexError:
                        pass
                    
                    summ_usl = Decimal(usl.getElementsByTagName("SUMV_USL")[0].childNodes[0].data)
                    if summ_usl != 0:
                        self.consolidated_ks['summ'] += summ_usl
                        self.consolidated_ks['pacients'] += 1
    
    def __consolidated_bill_ds(self, cases, settings):
        """
        Сбор данных для сводного счета по дневному стационару
        """
        
        for case in cases:
            nsvod = int(case.getElementsByTagName("NSVOD")[0].childNodes[0].data)
            if nsvod//100 == settings.priznak_ds:
                self.consolidated_ds['amount_of_days'] += int(case.getElementsByTagName("KD_Z")[0].childNodes[0].data)
                for usl in case.getElementsByTagName("USL"):
                    try:
                        self.consolidated_ds['amount_of_days_paid'] += int(usl.getElementsByTagName("KD")[0].childNodes[0].data)
                    except IndexError:
                        pass
                    
                    summ_usl = Decimal(usl.getElementsByTagName("SUMV_USL")[0].childNodes[0].data)
                    if summ_usl != 0:
                        self.consolidated_ds['summ'] += summ_usl
                        self.consolidated_ds['pacients'] += 1

    def __consolidated_bill_ks_insurance_company(self, cases, settings):
        """
        Сбор данных для сводного счета по круглосуточному стационару
        """
        for case in cases:
            nsvod = int(case.getElementsByTagName("NSVOD")[0].childNodes[0].data)
            if nsvod//100 == settings.priznak_ks:
                for usl in case.getElementsByTagName("USL"):
                    consolidated_ks_insurance_company = {'kod_lpu': None,             # 0 Код ЛПУ оказания
                                                         'podr': None,                # 1 Код подразделения оказания помощи
                                                         'pacients': 0,               # 3 Колчисевто выбывших пациентов
                                                         'amount_of_days': None,      # 4 Количество койкодней фактическое
                                                         'amount_of_days_paid': None, # 5 Количество койкодней оплачено
                                                         'fksg': None,                # 2 Код примененного КСГ, поле code_usl
                                                         'summ_usl': 0}               # 6 Сумма за Услугу
                    try:
                        kd = int(usl.getElementsByTagName("KD")[0].childNodes[0].data)
                        consolidated_ks_insurance_company['amount_of_days'] = kd
                        consolidated_ks_insurance_company['amount_of_days_paid'] = kd
                    except IndexError:
                        pass
                    summ_usl = Decimal(usl.getElementsByTagName("SUMV_USL")[0].childNodes[0].data)
                    if consolidated_ks_insurance_company['amount_of_days'] != None or summ_usl != Decimal(0):
                        consolidated_ks_insurance_company['kod_lpu'] = int(usl.getElementsByTagName("KODLPU")[0].childNodes[0].data)
                        consolidated_ks_insurance_company['podr'] = int(usl.getElementsByTagName("PODR")[0].childNodes[0].data)
                        consolidated_ks_insurance_company['fksg'] = usl.getElementsByTagName("CODE_USL")[0].childNodes[0].data
                        consolidated_ks_insurance_company['summ_usl'] = summ_usl
                        consolidated_ks_insurance_company['pacients'] = 1 if summ_usl != 0 else 0
                        self.__consolidated_ks_add_finded_case(consolidated_ks_insurance_company)

    def __consolidated_ks_add_finded_case(self, consolidated_ks_insurance_company):
        """
        Добавляем информацию о найденном случае в общий словарь круглосуточного стационара
        """
        address_cell_kod_lpu = -1
        address_cell_podr = -1
        address_cell_fksg = -1
        
        # Ищем ЛПУ в словаре
        if consolidated_ks_insurance_company['kod_lpu'] in self.consolidated_ks_insurance_company['kod_lpu']:
            # ЛПУ есть в словаре, ищем позицию
            for cell_kod_lpu in range(len(self.consolidated_ks_insurance_company['kod_lpu'])):
                if self.consolidated_ks_insurance_company['kod_lpu'][cell_kod_lpu] == consolidated_ks_insurance_company['kod_lpu']:
                    address_cell_kod_lpu = cell_kod_lpu
                    break
        else:
            # ЛПУ нет в словаре, добавляем и запоминаем позицию
            address_cell_kod_lpu = self.__consolidated_ks_insurance_company_add_new_lpu(consolidated_ks_insurance_company['kod_lpu'], consolidated_ks_insurance_company['podr'], consolidated_ks_insurance_company['fksg'])
        
        # Ищем профиль в словаре
        if address_cell_podr == -1 and (consolidated_ks_insurance_company['podr'] in self.consolidated_ks_insurance_company['podr'][address_cell_kod_lpu]):
            # профиль есть в списке, ищем позицию
            for cell_podr in range(0, len(self.consolidated_ks_insurance_company['podr'][address_cell_kod_lpu])):
                if self.consolidated_ks_insurance_company['podr'][address_cell_kod_lpu][cell_podr] == consolidated_ks_insurance_company['podr']:
                    address_cell_podr = cell_podr
                    break
        else:
            # профиля нет в списке, добавляем
            address_cell_podr = self.__consolidated_ks_insurance_company_add_new_podr(consolidated_ks_insurance_company['podr'], consolidated_ks_insurance_company['fksg'], address_cell_kod_lpu)
        
        # Ищем код услуги в списке
        if address_cell_fksg == -1 and consolidated_ks_insurance_company['fksg'] in self.consolidated_ks_insurance_company['fksg'][address_cell_kod_lpu][address_cell_podr]:
            # код услуги есть в списке, ищем позицию
            for cell_fksg in range(len(self.consolidated_ks_insurance_company['fksg'][address_cell_kod_lpu][address_cell_podr])):
                if self.consolidated_ks_insurance_company['fksg'][address_cell_kod_lpu][address_cell_podr][cell_fksg] == consolidated_ks_insurance_company['fksg']:
                    address_cell_fksg = cell_fksg
                    break
        else:
            # кода услуги нет в списке, добавляем
            address_cell_fksg = self.__consolidated_ks_insurance_company_add_new_fksg(consolidated_ks_insurance_company['fksg'], address_cell_kod_lpu, address_cell_podr)
        # прибавляем данные для лп, профиля, кода услуги
        self.consolidated_ks_insurance_company['pacients'][address_cell_kod_lpu][address_cell_podr][address_cell_fksg] += consolidated_ks_insurance_company['pacients']
        self.consolidated_ks_insurance_company['amount_of_days'][address_cell_kod_lpu][address_cell_podr][address_cell_fksg] += consolidated_ks_insurance_company['amount_of_days']
        self.consolidated_ks_insurance_company['amount_of_days_paid'][address_cell_kod_lpu][address_cell_podr][address_cell_fksg] += consolidated_ks_insurance_company['amount_of_days_paid']
        self.consolidated_ks_insurance_company['summ'][address_cell_kod_lpu][address_cell_podr][address_cell_fksg] += consolidated_ks_insurance_company['summ_usl']
    
    def __consolidated_ks_insurance_company_add_new_lpu(self, kod_lpu, podr, fksg):
        """
        Добавляем новую больницу в общий словарь круглосуточного стационара
        """
        self.consolidated_ks_insurance_company['kod_lpu'].append(kod_lpu)
        address_cell_kod_lpu = len(self.consolidated_ks_insurance_company['kod_lpu']) - 1
        self.consolidated_ks_insurance_company['podr'].append([])
        self.consolidated_ks_insurance_company['fksg'].append([])
        self.consolidated_ks_insurance_company['pacients'].append([])
        self.consolidated_ks_insurance_company['amount_of_days'].append([])
        self.consolidated_ks_insurance_company['amount_of_days_paid'].append([])
        self.consolidated_ks_insurance_company['summ'].append([])
        self.__consolidated_ks_insurance_company_add_new_podr(podr, fksg, address_cell_kod_lpu)
        return address_cell_kod_lpu
    
    def __consolidated_ks_insurance_company_add_new_podr(self, podr, fksg, address_cell_kod_lpu):
        """
        Добавляем новое подразделение в общий словарь круглосуточного стационара
        """
        self.consolidated_ks_insurance_company['podr'][address_cell_kod_lpu].append(podr)
        address_cell_podr = len(self.consolidated_ks_insurance_company['podr'][address_cell_kod_lpu]) - 1
        self.consolidated_ks_insurance_company['fksg'][address_cell_kod_lpu].append([])
        self.consolidated_ks_insurance_company['pacients'][address_cell_kod_lpu].append([])
        self.consolidated_ks_insurance_company['amount_of_days'][address_cell_kod_lpu].append([])
        self.consolidated_ks_insurance_company['amount_of_days_paid'][address_cell_kod_lpu].append([])
        self.consolidated_ks_insurance_company['summ'][address_cell_kod_lpu].append([])
        self.__consolidated_ks_insurance_company_add_new_fksg(fksg, address_cell_kod_lpu, address_cell_podr)
        return address_cell_podr
    
    def __consolidated_ks_insurance_company_add_new_fksg(self, fksg, address_cell_kod_lpu, address_cell_podr):
        """
        Добавляем новую услугу в подразделение в общий словарь круглосуточного стационара
        """
        self.consolidated_ks_insurance_company['fksg'][address_cell_kod_lpu][address_cell_podr].append(fksg)
        self.consolidated_ks_insurance_company['pacients'][address_cell_kod_lpu][address_cell_podr].append(0)
        self.consolidated_ks_insurance_company['amount_of_days'][address_cell_kod_lpu][address_cell_podr].append(0)
        self.consolidated_ks_insurance_company['amount_of_days_paid'][address_cell_kod_lpu][address_cell_podr].append(0)
        self.consolidated_ks_insurance_company['summ'][address_cell_kod_lpu][address_cell_podr].append(0)
        return len(self.consolidated_ks_insurance_company['fksg'][address_cell_kod_lpu][address_cell_podr]) - 1

    def __consolidated_bill_ds_insurance_company(self, cases, settings):
        """
        Сбор данных для сводного счета по дневному стационару
        """
        for case in cases:
            nsvod = int(case.getElementsByTagName("NSVOD")[0].childNodes[0].data)
            if nsvod//100 == settings.priznak_ds:
                for usl in case.getElementsByTagName("USL"):
                    consolidated_ds_insurance_company = {'kod_lpu': None,             # 0 Код ЛПУ оказания
                                                         'podr': None,                # 1 Код подразделения оказания помощи
                                                         'pacients': 0,               # 3 Колчисевто выбывших пациентов
                                                         'amount_of_days': None,      # 4 Количество койкодней фактическое
                                                         'amount_of_days_paid': None, # 5 Количество койкодней оплачено
                                                         'fksg': None,                # 2 Код примененного КСГ, поле code_usl
                                                         'summ_usl': 0}               # 6 Сумма за Услугу
                    try:
                        kd = int(usl.getElementsByTagName("KD")[0].childNodes[0].data)
                        consolidated_ds_insurance_company['amount_of_days'] = kd
                        consolidated_ds_insurance_company['amount_of_days_paid'] = kd
                    except IndexError:
                        pass
                    summ_usl = Decimal(usl.getElementsByTagName("SUMV_USL")[0].childNodes[0].data)
                    if consolidated_ds_insurance_company['amount_of_days'] != None or summ_usl != Decimal(0):
                        consolidated_ds_insurance_company['kod_lpu'] = int(usl.getElementsByTagName("KODLPU")[0].childNodes[0].data)
                        consolidated_ds_insurance_company['podr'] = int(usl.getElementsByTagName("PODR")[0].childNodes[0].data)
                        consolidated_ds_insurance_company['fksg'] = usl.getElementsByTagName("CODE_USL")[0].childNodes[0].data
                        consolidated_ds_insurance_company['summ_usl'] = summ_usl
                        consolidated_ds_insurance_company['pacients'] = 1 if summ_usl != 0 else 0
                        self.__consolidated_ds_add_finded_case(consolidated_ds_insurance_company)

    def __consolidated_ds_add_finded_case(self, consolidated_ds_insurance_company):
        """
        Добавляем информацию о найденном случае в общий словарь дневного стационара
        """
        address_cell_kod_lpu = -1
        address_cell_podr = -1
        address_cell_fksg = -1
        
        # Ищем ЛПУ в словаре
        if consolidated_ds_insurance_company['kod_lpu'] in self.consolidated_ds_insurance_company['kod_lpu']:
            # ЛПУ есть в словаре, ищем позицию
            for cell_kod_lpu in range(len(self.consolidated_ds_insurance_company['kod_lpu'])):
                if self.consolidated_ds_insurance_company['kod_lpu'][cell_kod_lpu] == consolidated_ds_insurance_company['kod_lpu']:
                    address_cell_kod_lpu = cell_kod_lpu
                    break
        else:
            # ЛПУ нет в словаре, добавляем и запоминаем позицию
            address_cell_kod_lpu = self.__consolidated_ds_insurance_company_add_new_lpu(consolidated_ds_insurance_company['kod_lpu'], consolidated_ds_insurance_company['podr'], consolidated_ds_insurance_company['fksg'])
        
        # Ищем профиль в словаре
        if address_cell_podr == -1 and (consolidated_ds_insurance_company['podr'] in self.consolidated_ds_insurance_company['podr'][address_cell_kod_lpu]):
            # профиль есть в списке, ищем позицию
            for cell_podr in range(0, len(self.consolidated_ds_insurance_company['podr'][address_cell_kod_lpu])):
                if self.consolidated_ds_insurance_company['podr'][address_cell_kod_lpu][cell_podr] == consolidated_ds_insurance_company['podr']:
                    address_cell_podr = cell_podr
                    break
        else:
            # профиля нет в списке, добавляем
            address_cell_podr = self.__consolidated_ds_insurance_company_add_new_podr(consolidated_ds_insurance_company['podr'], consolidated_ds_insurance_company['fksg'], address_cell_kod_lpu)
        
        # Ищем код услуги в списке
        if address_cell_fksg == -1 and consolidated_ds_insurance_company['fksg'] in self.consolidated_ds_insurance_company['fksg'][address_cell_kod_lpu][address_cell_podr]:
            # код услуги есть в списке, ищем позицию
            for cell_fksg in range(len(self.consolidated_ds_insurance_company['fksg'][address_cell_kod_lpu][address_cell_podr])):
                if self.consolidated_ds_insurance_company['fksg'][address_cell_kod_lpu][address_cell_podr][cell_fksg] == consolidated_ds_insurance_company['fksg']:
                    address_cell_fksg = cell_fksg
                    break
        else:
            # кода услуги нет в списке, добавляем
            address_cell_fksg = self.__consolidated_ds_insurance_company_add_new_fksg(consolidated_ds_insurance_company['fksg'], address_cell_kod_lpu, address_cell_podr)
        # прибавляем данные для лп, профиля, кода услуги
        self.consolidated_ds_insurance_company['pacients'][address_cell_kod_lpu][address_cell_podr][address_cell_fksg] += consolidated_ds_insurance_company['pacients']
        self.consolidated_ds_insurance_company['amount_of_days'][address_cell_kod_lpu][address_cell_podr][address_cell_fksg] += consolidated_ds_insurance_company['amount_of_days']
        self.consolidated_ds_insurance_company['amount_of_days_paid'][address_cell_kod_lpu][address_cell_podr][address_cell_fksg] += consolidated_ds_insurance_company['amount_of_days_paid']
        self.consolidated_ds_insurance_company['summ'][address_cell_kod_lpu][address_cell_podr][address_cell_fksg] += consolidated_ds_insurance_company['summ_usl']
    
    def __consolidated_ds_insurance_company_add_new_lpu(self, kod_lpu, podr, fksg):
        """
        Добавляем новую больницу в общий словарь дневного стационара
        """
        self.consolidated_ds_insurance_company['kod_lpu'].append(kod_lpu)
        address_cell_kod_lpu = len(self.consolidated_ds_insurance_company['kod_lpu']) - 1
        self.consolidated_ds_insurance_company['podr'].append([])
        self.consolidated_ds_insurance_company['fksg'].append([])
        self.consolidated_ds_insurance_company['pacients'].append([])
        self.consolidated_ds_insurance_company['amount_of_days'].append([])
        self.consolidated_ds_insurance_company['amount_of_days_paid'].append([])
        self.consolidated_ds_insurance_company['summ'].append([])
        self.__consolidated_ds_insurance_company_add_new_podr(podr, fksg, address_cell_kod_lpu)
        return address_cell_kod_lpu
    
    def __consolidated_ds_insurance_company_add_new_podr(self, podr, fksg, address_cell_kod_lpu):
        """
        Добавляем новое подразделение в общий словарь дневного стационара
        """
        self.consolidated_ds_insurance_company['podr'][address_cell_kod_lpu].append(podr)
        address_cell_podr = len(self.consolidated_ds_insurance_company['podr'][address_cell_kod_lpu]) - 1
        self.consolidated_ds_insurance_company['fksg'][address_cell_kod_lpu].append([])
        self.consolidated_ds_insurance_company['pacients'][address_cell_kod_lpu].append([])
        self.consolidated_ds_insurance_company['amount_of_days'][address_cell_kod_lpu].append([])
        self.consolidated_ds_insurance_company['amount_of_days_paid'][address_cell_kod_lpu].append([])
        self.consolidated_ds_insurance_company['summ'][address_cell_kod_lpu].append([])
        self.__consolidated_ds_insurance_company_add_new_fksg(fksg, address_cell_kod_lpu, address_cell_podr)
        return address_cell_podr
    
    def __consolidated_ds_insurance_company_add_new_fksg(self, fksg, address_cell_kod_lpu, address_cell_podr):
        """
        Добавляем новую услугу в подразделение в общий словарь дневного стационара
        """
        self.consolidated_ds_insurance_company['fksg'][address_cell_kod_lpu][address_cell_podr].append(fksg)
        self.consolidated_ds_insurance_company['pacients'][address_cell_kod_lpu][address_cell_podr].append(0)
        self.consolidated_ds_insurance_company['amount_of_days'][address_cell_kod_lpu][address_cell_podr].append(0)
        self.consolidated_ds_insurance_company['amount_of_days_paid'][address_cell_kod_lpu][address_cell_podr].append(0)
        self.consolidated_ds_insurance_company['summ'][address_cell_kod_lpu][address_cell_podr].append(0)
        return len(self.consolidated_ds_insurance_company['fksg'][address_cell_kod_lpu][address_cell_podr]) - 1

    def __capitalize(self):
        """
        Сума прописью
        """
        summ = (self.bill_data['ks_current_month_summ'] + self.bill_data['ks_previous_month_summ'] + 
                self.bill_data['ds_current_month_summ'] + self.bill_data['ds_previous_month_summ'] +
                self.bill_data['ambulance_current_month_summ'] + self.bill_data['ambulance_previous_month_summ'] +
                self.bill_data['smp_current_month_summ'] + self.bill_data['smp_previous_month_summ'] +
                self.bill_data['summ_pf'] + self.bill_data['summ_fap'])
        value = pytils.numeral.rubles (int(summ))
        #Отделяем рубли от копеек
        penny = int ((summ * 100 - int (summ ) * 100))
        self.bill_data['capitalize'] = value.capitalize() + ' ' + str (penny) + ' ' + self.__capitalize_penny (penny)
    
    def __capitalize_penny(self, penny):
        """
        Возвращает значенеи текстовой подписи для копеек
        """
        #         0         1         2         3         4         5
        capitalize_penny = ['копеек','копейка','копейки','копейки','копейки','копеек','копеек','копеек','копеек','копеек']
        if penny < 20 and penny > 10:
            return 'копеек'
        else:
            up_to_ten_penny = int(penny / 10) * 10
            index_capitalize_penny = penny - up_to_ten_penny
            return capitalize_penny[index_capitalize_penny]

    def __select_months(self):
        """
        Заполняем поля текущего и проошедшего месяца
        """
        pass


class FormatGenerator:
    """
    Базовый класс 
    """
    start_row = ''
    current_row = ''
    cell_final_summ = ''
    dict_cells_ambulance_preresult = {'visits': [],
                            'appeal': [],
                            'services_in_sluch': [],
                            'dentistry_uet': [],
                            'individual_bill': [],
                            'summ_sluch': []}
    dict_cells_hospital_preresult = {'pacients': [],
                                     'amount_of_days': [],
                                     'amount_of_days_paid': [],
                                     'summ': []}
    dict_cells_hospital_subtotal = {'pacients': [],
                                     'amount_of_days': [],
                                     'amount_of_days_paid': [],
                                     'summ': []}
    
    def __init__(self, *args, **kwargs):
        """
        
        """
        out_file = openpyxl.Workbook()
        self.out_sheet = out_file.active
        self.out_sheet.title = 'Лист'
        out_file.save('файл' + '.xlsx')
        out_file.close()

    def _select_format_for_smo(self, path_to_format_file_xml):
        """
        Возвращает записи из файла xml с форматом сводного счета
        """
        try:
            format_file = xml.dom.minidom.parse(path_to_format_file_xml)
            return format_file.getElementsByTagName("cells")[0]
        except IndexError:
            print ('Файл не содержит данных о формате сводного счета')
            sys.exit(0)
        print('Не найдена форма сводного счета')
        sys.exit(0)

    def _bill_formation(self, format_for_smo, medical_organization_data, insurance_company_data, data):
        """
        Заполняем счет
        """
        cells_format = format_for_smo.getElementsByTagName("cell")
        for cell_format in cells_format:
            try:
                first_cell = cell_format.getElementsByTagName("first_cell")[0].childNodes[0].data
            except IndexError:
                print('Файл формы поврежден, не указаны first_cell')
                sys.exit(1)
            # Обрамление ячейки
            formation_border = self.__cell_border(cell_format)
            self.out_sheet[first_cell].border = formation_border
            merge_to_cell = None
            for merged_cell in cell_format.getElementsByTagName("mcell"):
                merge_to_cell = merged_cell.childNodes[0].data
                self.out_sheet[merge_to_cell].border = formation_border
            # Объединение ячеек
            if merge_to_cell != None:
                cells_to_merge = first_cell + ':' + merge_to_cell
                self.out_sheet.merge_cells(cells_to_merge)
            # Шрифт
            self.out_sheet[first_cell].font = self.__cell_font(cell_format)
            # Выравнивание текста в ячейке
            self.out_sheet[first_cell].alignment = self.__cell_alignment(cell_format)
            # Формат вывода сумм
            try:
                self.out_sheet[first_cell].number_format = cell_format.getElementsByTagName("number_format")[0].childNodes[0].data
            except IndexError:
                pass
            # Запись значения
            try:
                value_for_cell = cell_format.getElementsByTagName("value")[0].childNodes[0].data
            except IndexError:
                value_for_cell = ''
            if value_for_cell.find('var_') != -1:
                value_for_cell = self._change_value_if_var(value_for_cell, medical_organization_data, insurance_company_data, data)
            self.out_sheet[first_cell].value = value_for_cell
        # Ширина столбцов
        self.__column_format(format_for_smo)
        # Высота столбцов
        self.__row_format(format_for_smo)
        # Параметры страницы
        self.__page_setup(format_for_smo)
    
    def _change_value_if_var(self, value_for_cell, medical_organization_data, insurance_company_data, data):
        return value_for_cell
    
    def _change_value_consolidated_if_var(self, value, consolidated_insurance_company, index_lpu, index_profil, index_kod_usl, names_profiles):
        """
        Получает данные для сводного счета и название перевенной,
        возвращает значение переменной
        """
        value = value[17:]
        if value == 'kod_lpu':
            return consolidated_insurance_company[value][index_lpu]
        # Поликлинические профили
        elif value == 'profil' or value == 'podr':
            try:
                field = names_profiles[str(consolidated_insurance_company[value][index_lpu][index_profil])[:4]]
            except KeyError:
                field = consolidated_insurance_company[value][index_lpu][index_profil]
            return field
        elif value == 'number_value':
            return consolidated_insurance_company['number_value']
        else:
            return consolidated_insurance_company[value][index_lpu][index_profil][index_kod_usl]

        
        return value
    
    def __cell_border(self, cell_format):
        """
        Зашружаем оформление границ ячейки
        возвращаем 
        """
        border_dict = {'left': None,
                   'right': None,
                   'top': None,
                   'bottom': None}
        try:
            border_dict['left'] = cell_format.getElementsByTagName("left")[0].childNodes[0].data
        except IndexError:
            pass
        try:
            border_dict['right'] = cell_format.getElementsByTagName("right")[0].childNodes[0].data
        except IndexError:
            pass
        try:
            border_dict['top'] = cell_format.getElementsByTagName("top")[0].childNodes[0].data
        except IndexError:
            pass
        try:
            border_dict['bottom'] = cell_format.getElementsByTagName("bottom")[0].childNodes[0].data
        except IndexError:
            pass
        
        formation_border = openpyxl.styles.Border()
        if border_dict['left']!=None:
            formation_border.left=openpyxl.styles.Side(border_style='thin', color=border_dict['left'])
        if border_dict['right']!=None:
            formation_border.right=openpyxl.styles.Side(border_style='thin', color=border_dict['right'])
        if border_dict['top']!=None:
            formation_border.top=openpyxl.styles.Side(border_style='thin',color=border_dict['top'])
        if border_dict['bottom']!=None:
            formation_border.bottom=openpyxl.styles.Side(border_style='thin',color=border_dict['bottom'])
        return formation_border

    def __cell_font(self, cell_format):
        """
        Загружаем данные шрифта, его размера и стилей
        возвращаем стиль шрифта
        """
        font = {'name': '',
                'size': float(0),
                'bold': int(0),
                'italic': int(0)}
        try:
            font['name'] = str(cell_format.getElementsByTagName("font_name")[0].childNodes[0].data)
        except IndexError:
            print('Нет данных о названии шрифта для ячейки: ', cell_format.getElementsByTagName("first_cell")[0].childNodes[0].data)
            sys.exit(1)
        try:
            font['size'] = float(cell_format.getElementsByTagName("size")[0].childNodes[0].data)
        except IndexError:
            print('Нет данных о размере шрифта для ячейки: ', cell_format.getElementsByTagName("first_cell")[0].childNodes[0].data)
            sys.exit(1)
        try:
            font['bold'] = int(cell_format.getElementsByTagName("bold")[0].childNodes[0].data)
        except IndexError:
            print('Нет данных о жирности шрифта для ячейки: ', cell_format.getElementsByTagName("first_cell")[0].childNodes[0].data)
            sys.exit(1)
        try:
            font['italic'] = int(cell_format.getElementsByTagName("italic")[0].childNodes[0].data)
        except IndexError:
            print('Нет данных о курсивности шрифта для ячейки: ', cell_format.getElementsByTagName("first_cell")[0].childNodes[0].data)
            sys.exit(1)
        font_formation = openpyxl.styles.Font(name=font['name'],
                                              size=font['size'],
                                              bold=font['bold'],
                                              italic=font['italic'])
        return font_formation

    def __cell_alignment(self, cell_format):
        """
        Загружаем данные по выравниванию в ячейке,
        возвращаем стиль выравниваняи
        """
        alignment = {'horizontal': False,
                    'vertical': False,
                    'wrap_text': False}
        try:
            alignment['horizontal'] = str(cell_format.getElementsByTagName("horizontal")[0].childNodes[0].data)
        except IndexError:
            print('Не указано выравние по горизонтали для ячейки: ', cell_format.getElementsByTagName("first_cell")[0].childNodes[0].data)
            sys.exit(1)
        try:
            alignment['vertical'] = str(cell_format.getElementsByTagName("vertical")[0].childNodes[0].data)
        except IndexError:
            print('Не указано выравние по вертикале для ячейки: ', cell_format.getElementsByTagName("first_cell")[0].childNodes[0].data)
            sys.exit(1)
        try:
            alignment['wrap_text'] = int(cell_format.getElementsByTagName("wrap_text")[0].childNodes[0].data)
        except IndexError:
            print('Не указано требование к многострочности в ячейке : ', cell_format.getElementsByTagName("first_cell")[0].childNodes[0].data)
            sys.exit(1)
        return openpyxl.styles.Alignment(horizontal=alignment['horizontal'],
                                         vertical=alignment['vertical'],
                                         wrap_text=alignment['wrap_text'])

    def __column_format(self, format_for_smo):
        """
        Ширина и видимость столбцов
        """
        for column in format_for_smo.getElementsByTagName("column"):
            column_letter = column.getElementsByTagName("letter")[0].childNodes[0].data
            column_width = float(column.getElementsByTagName("width")[0].childNodes[0].data)
            column_hidden = int(column.getElementsByTagName("hidden")[0].childNodes[0].data)
            self.out_sheet.column_dimensions[column_letter].width = column_width * 5.0987
            self.out_sheet.column_dimensions[column_letter].hidden = column_hidden

    def __row_format(self, format_for_smo):
        """
        Выоста строк и их видимость
        """
        for row in format_for_smo.getElementsByTagName("row"):
            row_number = int(row.getElementsByTagName("num")[0].childNodes[0].data)
            row_height = float(row.getElementsByTagName("height")[0].childNodes[0].data)
            row_hidden = int(row.getElementsByTagName("hidden")[0].childNodes[0].data)
            self.out_sheet.row_dimensions[row_number].height = row_height * 28.3571
            self.out_sheet.row_dimensions[row_number].hidden = row_hidden

    def __page_setup(self, format_for_smo):
        """
        Отступы, масштаб страницы и формат бумаги
        """
        page_margin_left = float(format_for_smo.getElementsByTagName("page_left")[0].childNodes[0].data)
        page_margin_right = float(format_for_smo.getElementsByTagName("page_right")[0].childNodes[0].data)
        page_margin_top = float(format_for_smo.getElementsByTagName("page_top")[0].childNodes[0].data)
        page_margin_bottom = float(format_for_smo.getElementsByTagName("page_bottom")[0].childNodes[0].data)
        page_scale = int(format_for_smo.getElementsByTagName("page_scale")[0].childNodes[0].data)
        page_orientation = format_for_smo.getElementsByTagName("orientation")[0].childNodes[0].data
        
        self.out_sheet.page_margins.left = page_margin_left / 2.54
        self.out_sheet.page_margins.right = page_margin_right / 2.54
        self.out_sheet.page_margins.top = page_margin_top / 2.54
        self.out_sheet.page_margins.bottom = page_margin_bottom / 2.54
        self.out_sheet.page_setup.scale = page_scale
        self.out_sheet.page_setup.paperSize = '9'
        self.out_sheet.page_setup.orientation = page_orientation

    def _format_cell_consolidated(self, cell_format):
        """
        Формат ячейки для сводного счета с динамическим количством строк
        """
        # Первая ячейка
        first_cell = cell_format.getElementsByTagName("first_cell")[0].childNodes[0].data + str(self.current_row)
        # Обрамление ячейки
        formation_border = self.__cell_border(cell_format)
        self.out_sheet[first_cell].border = formation_border
        merge_to_cell = None
        for merged_cell in cell_format.getElementsByTagName("mcell"):
            merge_to_cell = merged_cell.childNodes[0].data + str(self.current_row)
            self.out_sheet[merge_to_cell].border = formation_border
        # Объединение ячеек
        if merge_to_cell != None:
            cells_to_merge = first_cell + ':' + merge_to_cell
            self.out_sheet.merge_cells(cells_to_merge)
        # Шрифт
        self.out_sheet[first_cell].font = self.__cell_font(cell_format)
        # Выравнивание текста в ячейке
        self.out_sheet[first_cell].alignment = self.__cell_alignment(cell_format)
        # Формат вывода сумм
        try:
            self.out_sheet[first_cell].number_format = cell_format.getElementsByTagName("number_format")[0].childNodes[0].data
        except IndexError:
            pass
        return first_cell

    def _change_value_consolidated_ambulance_preresult_if_var(self, value, column, lpu):
        """
        Возвращет формулу для предварительного результата
        """
        if value.find('_summ_') != -1:
            value = value[9:]
            cell = column + str(self.current_row)
            self.dict_cells_ambulance_preresult[value].append(cell)
            cell = '=sum(' + column + str(self.start_row) + ':' + column + str(self.current_row - 1) + ')'
            return cell
        if value.find('kod_lpu_and_') != -1:
            value = value[16:] + str(lpu)
        return value
    
    def _change_value_consolidated_ambulance_result_if_var(self, value, column):
        """
        возвращает формулу для подведения итога
        """
        if value.find('final_summ_sluch') != -1:
            self.cell_final_summ = column + str(self.current_row)
        if value.find('var_summ_final_') != -1:
            value = value[15:]
            len_list_value = len(self.dict_cells_ambulance_preresult[value])
            cell = '='
            for index_value in range(len_list_value):
                if self.dict_cells_ambulance_preresult[value][len_list_value-1] != self.dict_cells_ambulance_preresult[value][index_value]:
                    cell += self.dict_cells_ambulance_preresult[value][index_value] + '+'
                else:
                    cell += self.dict_cells_ambulance_preresult[value][index_value]
            return cell
        return value

    def _change_value_consolidated_hospital_preresult_if_var(self, value, column, podr):
        """
        Возвращет формулу для предварительного результата
        """
        if value.find('_summ_') != -1:
            value = value[9:]
            cell = column + str(self.current_row)
            # Запоминаем ячейку с предварительным результатом по отделению
            self.dict_cells_hospital_preresult[value].append(cell)
            cell = '=sum(' + column + str(self.start_row) + ':' + column + str(self.current_row - 1) + ')'
            return cell
        if value.find('var_podr_and_') != -1:
            value = value[13:] + str(podr)
        return value

    def _change_value_consolidated_hosppital_subtotal_if_var(self, value, column, lpu):
        """
        возвращает формулу для подведения итога
        """
        if value.find('var_lpu_and_') != -1:
            return value[12:] + str(lpu)
        if value.find('var_summ_subtotal_') != -1:
            value = value[18:]
            len_list_value = len(self.dict_cells_hospital_preresult[value])
            cell = '='
            for index_value in range(len_list_value):
                if self.dict_cells_hospital_preresult[value][len_list_value-1] != self.dict_cells_hospital_preresult[value][index_value]:
                    cell += self.dict_cells_hospital_preresult[value][index_value] + '+'
                else:
                    cell += self.dict_cells_hospital_preresult[value][index_value]
            # Запоминаем ячейку с предварительным результатом по ЛПУ
            self.dict_cells_hospital_subtotal[value].append(column + str(self.current_row))
            return cell
        return value
    
    def _change_value_consolidated_hospital_result_if_var(self, value, column):
        """
        возвращает формулу для подведения итога
        """
        if value.find('var_summ_final_summ') != -1:
            self.cell_final_summ = column + str(self.current_row)
        if value.find('var_summ_final_') != -1:
            value = value[15:]
            len_list_value = len(self.dict_cells_hospital_subtotal[value])
            cell = '='
            for index_value in range(len_list_value):
                if self.dict_cells_hospital_subtotal[value][len_list_value-1] != self.dict_cells_hospital_subtotal[value][index_value]:
                    cell += self.dict_cells_hospital_subtotal[value][index_value] + '+'
                else:
                    cell += self.dict_cells_hospital_subtotal[value][index_value]
            return cell
        return value

    def _clear_dict_ambulance(self):
        self.dict_cells_ambulance_preresult = {'visits': [],
                                               'appeal': [],
                                               'services_in_sluch': [],
                                               'dentistry_uet': [],
                                               'individual_bill': [],
                                               'summ_sluch': []}
    
    def _clear_dict_hospital(self):
        self.dict_cells_hospital_preresult = {'pacients': [],
                                              'amount_of_days': [],
                                              'amount_of_days_paid': [],
                                              'summ': []}
    
    def _clear_dict_hospital_subtotal(self):
        self.dict_cells_hospital_subtotal = {'pacients': [],
                                             'amount_of_days': [],
                                             'amount_of_days_paid': [],
                                             'summ': []}


class BillGenerator(FormatGenerator):
    """
    Формирование счетов в страховые компании
    в формате файлов электронных таблиц
    """
    def __init__(self, *args, **kwargs):
        """
        Инициализация отчетов
        """
        try:
            settings = kwargs['settings']
        except KeyError:
            print('Для формирования счета необходимо передать настройки')
            sys.exit(1)
        try:
            bill_data = kwargs['bill_data']
        except KeyError:
            print('Не получены данные для формирования счета')
            sys.exit(1)
        format_for_smo = self.__select_format_for_smo(settings.bill_format_xls, bill_data['kod_smo'])
        
        bill_out_file = openpyxl.Workbook()
        self.out_sheet = bill_out_file.active
        self.out_sheet.title = 'Счет'
        super()._bill_formation(format_for_smo, settings.medical_organization_data,
                              settings.insurance_company_data,  bill_data)
        if not os.path.exists(settings.source_dir + '/счета'):
            os.mkdir(settings.source_dir + '/счета')
        bill_out_file.save(settings.source_dir + 'счета/' +
                           'Re_' + bill_data['number_bill'] + '.xlsx')
        bill_out_file.close()
        return None

    def __select_format_for_smo(self, path_to_format_file_xml, kod_smo):
        """
        Возвращает записи из файла xml с форматом счета
        для выбранной страховой компании
        """
        try:
            format_file = xml.dom.minidom.parse(path_to_format_file_xml)
            nodes_smo = format_file.getElementsByTagName("bill")
        except IndexError:
            print ('Файл не содержит данных о формате счета')
            sys.exit(0)
        try:
            for node_smo in nodes_smo:
                kods_smo_in_format = node_smo.getElementsByTagName("kod_smo")
                for kod_smo_in_format in kods_smo_in_format:
                    finded_kod_smo = kod_smo_in_format.childNodes[0].data
                    if finded_kod_smo.find(kod_smo) != -1:
                        return node_smo
        except IndexError:
            print ('Не найдены записи kod_smo')
            sys.exit(0)
        print('Не найдена форма счета для страховой компании: ', kod_smo)
        sys.exit(0)

    def _change_value_if_var(self, value, medical_organization_data, insurance_company_data, bill_data):
        """
        Получает название переменной. возвражает значение переменной и данных счета
        """
        var = value[3:]
        if var.find('_mo_') != -1:
            var = var[4:]
            return medical_organization_data[var]
        if var.find('_smo_') != -1:
            smo_list_index = insurance_company_data['kod_smo'].index(bill_data['kod_smo'])
            var = var[5:]
            return insurance_company_data[var][smo_list_index]
        if var.find('_bill_') != -1:
            var = var[6:]
            return bill_data[var]
        return value


class ConsolidatedAmbulanceBillInSmoGenerator(FormatGenerator):
    """
    Формирование поликлинического сводного счета по
    страховой компании в форме электронных таблиц
    """
    
    def __init__(self, *args, **kwargs):
        """
        Инициализация отчетов
        """
        self._clear_dict_ambulance()
        try:
            settings = kwargs['settings']
        except KeyError:
            print('Для формирования счета необходимо передать настройки')
            sys.exit(1)
        try:
            consolidated_insurance_company = kwargs['consolidated_insurance_company']
        except KeyError:
            print('Не получены данные для формирования сводного счета')
            sys.exit(1)
        format_for_smo = super()._select_format_for_smo(settings.consolidated_ambulance_insurance_company_format_xls)
        
        out_file = openpyxl.Workbook()
        self.out_sheet = out_file.active
        self.out_sheet.title = 'Сводный'
        super()._bill_formation(format_for_smo, settings.medical_organization_data,
                              settings.insurance_company_data,  consolidated_insurance_company)
        self.start_row = int(format_for_smo.getElementsByTagName("start_row")[0].childNodes[0].data)
        self._format_consolidated(format_for_smo, consolidated_insurance_company, settings.medical_organization_data, settings.insurance_company_data, settings.dict_ambulance_profiles)
        if not os.path.exists(settings.source_dir + '/поликлиника'):
            os.mkdir(settings.source_dir + '/поликлиника')
        out_file.save(settings.source_dir + 'поликлиника/' +
                           'сводный_' + consolidated_insurance_company['kod_smo'] + '.xlsx')
        out_file.close()
        return None
    
    def _change_value_if_var(self, value, medical_organization_data, insurance_company_data, consolidated_insurance_company):
        """
        Получает название переменной. возвражает значение переменной и данных счета
        """
        var = value[3:]
        if var.find('_mo_') != -1:
            var = var[4:]
            return medical_organization_data[var]
        if var.find('_smo_') != -1:
            smo_list_index = insurance_company_data['kod_smo'].index(consolidated_insurance_company['kod_smo'])
            var = var[5:]
            return insurance_company_data[var][smo_list_index]
        if var.find('_consolidated_') != -1:
            var = var[14:]
            return consolidated_insurance_company[var]
        return value
    
    def _format_consolidated(self, format_for_smo, consolidated_insurance_company, medical_organization_data, insurance_company_data, names_profiles):
        """
        Заполняем данные сводного счета по подразделениям,
        профилям и услугам
        """
        self.current_row = self.start_row
        for index_lpu in range(len(consolidated_insurance_company['kod_lpu'])):
            self.start_row = self.current_row
            for index_profil in range(len(consolidated_insurance_company['profil'][index_lpu])):
                for index_kod_usl in range(len(consolidated_insurance_company['kod_usl'][index_lpu][index_profil])):
                    for cell_format in format_for_smo.getElementsByTagName("rcell"):
                        first_cell = super()._format_cell_consolidated(cell_format)
                        # Запись значения
                        try:
                            value_for_cell = cell_format.getElementsByTagName("value")[0].childNodes[0].data
                        except IndexError:
                            value_for_cell = ''
                        if value_for_cell.find('var_') != -1:
                            value_for_cell = self._change_value_consolidated_if_var(value_for_cell, consolidated_insurance_company, index_lpu, index_profil, index_kod_usl, names_profiles)
                        self.out_sheet[first_cell].value = value_for_cell
                    self.current_row += 1
            # Предварительные результаты
            for cell_format in format_for_smo.getElementsByTagName("pecell"):
                first_cell = super()._format_cell_consolidated(cell_format)
                column = cell_format.getElementsByTagName("first_cell")[0].childNodes[0].data
                # Запись значения
                try:
                    value_for_cell = cell_format.getElementsByTagName("value")[0].childNodes[0].data
                except IndexError:
                    value_for_cell = ''
                if value_for_cell.find('var_') != -1:
                    value_for_cell = super()._change_value_consolidated_ambulance_preresult_if_var(value_for_cell, column, consolidated_insurance_company['kod_lpu'][index_lpu])
                self.out_sheet[first_cell].value = value_for_cell
            self.current_row += 1
        
        # Окончательные результаты
        for cell_format in format_for_smo.getElementsByTagName("ecell"):
            first_cell = super()._format_cell_consolidated(cell_format)
            # Запись значения
            try:
                value_for_cell = cell_format.getElementsByTagName("value")[0].childNodes[0].data
            except IndexError:
                value_for_cell = ''
            if value_for_cell.find('var_') != -1:
                value_for_cell = super()._change_value_consolidated_ambulance_result_if_var(value_for_cell, column)
            self.out_sheet[first_cell].value = value_for_cell
        self.current_row += 3
        # Завершающая строка
        for cell_format in format_for_smo.getElementsByTagName("edcell"):
            first_cell = super()._format_cell_consolidated(cell_format)
            # Запись значения
            try:
                value_for_cell = cell_format.getElementsByTagName("value")[0].childNodes[0].data
            except IndexError:
                value_for_cell = ''
            if value_for_cell.find('summ_final_cell') != -1:
                value_for_cell = '=' + self.cell_final_summ
            self.out_sheet[first_cell].value = value_for_cell


class ConsolidatedHospitalBillInSmoGenerator(FormatGenerator):
    """
    Формирование стационару (круглосуточный или дневной) сводного
    счета по страховой компании в форме электронных таблиц
    """
    def __init__(self, *args, **kwargs):
        self._clear_dict_hospital()
        self._clear_dict_hospital_subtotal()
        # Получаем настройки
        try:
            settings = kwargs['settings']
        except KeyError:
            print('Для формирования счета необходимо передать настройки')
            sys.exit(1)
        # Получаем данные сводного счета
        try:
            consolidated_insurance_company = kwargs['consolidated_insurance_company']
        except KeyError:
            print('Не получены данные для формирования сводного счета')
            sys.exit(1)
        format_for_smo = super()._select_format_for_smo(settings.consolidated_hospital_insurance_company_format_xls)
        
        if consolidated_insurance_company['type_hospital'] == settings.priznak_ks:
            consolidated_insurance_company['type_hospital'] = 'круглосуточного стационара'
            path_dir = '/стационар_круглосуточный/'
        elif consolidated_insurance_company['type_hospital'] == settings.priznak_ds:
            consolidated_insurance_company['type_hospital'] = 'дневного стационара'
            path_dir = '/стационар_дневной/'
        else:
            path_dir = '/стационар/'
        
        out_file = openpyxl.Workbook()
        self.out_sheet = out_file.active
        self.out_sheet.title = 'Сводный'
        super()._bill_formation(format_for_smo, settings.medical_organization_data,
                              settings.insurance_company_data,  consolidated_insurance_company)
        self.start_row = int(format_for_smo.getElementsByTagName("start_row")[0].childNodes[0].data)
        self._format_consolidated(format_for_smo, consolidated_insurance_company, settings.medical_organization_data, settings.insurance_company_data, settings.dict_hospital_profiles)
        if not os.path.exists(settings.source_dir + path_dir):
            os.mkdir(settings.source_dir + path_dir)
        out_file.save(settings.source_dir + path_dir +
                           'сводный_' + consolidated_insurance_company['kod_smo'] + '.xlsx')
        out_file.close()
        return None
    
    def _change_value_if_var(self, value, medical_organization_data, insurance_company_data, consolidated_insurance_company):
        """
        Получает название переменной. возвражает значение переменной и данных счета
        """
        var = value[3:]
        if var.find('_mo_') != -1:
            var = var[4:]
            return medical_organization_data[var]
        if var.find('_smo_') != -1:
            smo_list_index = insurance_company_data['kod_smo'].index(consolidated_insurance_company['kod_smo'])
            var = var[5:]
            return insurance_company_data[var][smo_list_index]
        if var.find('_consolidated_') != -1:
            var = var[14:]
            return str(consolidated_insurance_company[var])
        return value
    
    def _format_consolidated(self, format_for_smo, consolidated_insurance_company, medical_organization_data, insurance_company_data, names_profiles):
        """
        Заполняем данные сводного счета по подразделениям,
        профилям и услугам
        """
        self.current_row = self.start_row
        for index_lpu in range(len(consolidated_insurance_company['kod_lpu'])):
            for index_podr in range(len(consolidated_insurance_company['podr'][index_lpu])):
                self.start_row = self.current_row
                for index_fksg in range(len(consolidated_insurance_company['fksg'][index_lpu][index_podr])):
                    for cell_format in format_for_smo.getElementsByTagName("rcell"):
                        first_cell = super()._format_cell_consolidated(cell_format)
                        # Запись значения
                        try:
                            value_for_cell = cell_format.getElementsByTagName("value")[0].childNodes[0].data
                        except IndexError:
                            value_for_cell = ''
                        if value_for_cell.find('var_') != -1:
                            value_for_cell = self._change_value_consolidated_if_var(value_for_cell, consolidated_insurance_company, index_lpu, index_podr, index_fksg, names_profiles)
                        self.out_sheet[first_cell].value = value_for_cell
                    consolidated_insurance_company['number_value'] += 1
                    self.current_row += 1
                    
                # Предварительные результаты отделения
                for cell_format in format_for_smo.getElementsByTagName("pecell_podr"):
                    first_cell = super()._format_cell_consolidated(cell_format)
                    column = cell_format.getElementsByTagName("first_cell")[0].childNodes[0].data
                    # Запись значения
                    try:
                        value_for_cell = cell_format.getElementsByTagName("value")[0].childNodes[0].data
                    except IndexError:
                        value_for_cell = ''
                    if value_for_cell.find('var_') != -1:
                        value_for_cell = super()._change_value_consolidated_hospital_preresult_if_var(value_for_cell, column, names_profiles[str(consolidated_insurance_company['podr'][index_lpu][index_podr])[:4]])
                    self.out_sheet[first_cell].value = value_for_cell
                self.current_row += 1
                
            # Предварительные результаты ЛПУ
            for cell_format in format_for_smo.getElementsByTagName("pecell_lpu"):
                first_cell = super()._format_cell_consolidated(cell_format)
                column = cell_format.getElementsByTagName("first_cell")[0].childNodes[0].data
                # Запись значения
                try:
                    value_for_cell = cell_format.getElementsByTagName("value")[0].childNodes[0].data
                except IndexError:
                    value_for_cell = ''
                if value_for_cell.find('var_') != -1:
                    value_for_cell = super()._change_value_consolidated_hosppital_subtotal_if_var(value_for_cell, column, consolidated_insurance_company['kod_lpu'][index_lpu])
                self.out_sheet[first_cell].value = value_for_cell
            self.current_row += 1
            self._clear_dict_hospital()
        
        # Окончательные результаты
        for cell_format in format_for_smo.getElementsByTagName("ecell"):
            first_cell = super()._format_cell_consolidated(cell_format)
            # Запись значения
            try:
                value_for_cell = cell_format.getElementsByTagName("value")[0].childNodes[0].data
            except IndexError:
                value_for_cell = ''
            if value_for_cell.find('var_') != -1:
                value_for_cell = super()._change_value_consolidated_hospital_result_if_var(value_for_cell, column)
            self.out_sheet[first_cell].value = value_for_cell
        self.current_row += 3
        # Завершающая строка
        for cell_format in format_for_smo.getElementsByTagName("edcell"):
            first_cell = super()._format_cell_consolidated(cell_format)
            # Запись значения
            try:
                value_for_cell = cell_format.getElementsByTagName("value")[0].childNodes[0].data
            except IndexError:
                value_for_cell = ''
            if value_for_cell.find('summ_final_cell') != -1:
                value_for_cell = '=' + self.cell_final_summ
            self.out_sheet[first_cell].value = value_for_cell


class ConsolidatedAmbulanceBillGenerator(FormatGenerator):
    """
    Сводыные счета по поликлиникие с ТФОМС и без
    """
    def __init__(self, *args, **kwargs):
        """
        Инициализация отчетов
        """
        try:
            settings = kwargs['settings']
        except KeyError:
            print('Для формирования счета необходимо передать настройки')
            sys.exit(1)
        try:
            list_consolidated_smo = kwargs['list_consolidated_smo']
        except KeyError:
            print('Не получены данные для формирования сводного счета')
            sys.exit(1)
        self.format_for_smo = super()._select_format_for_smo(settings.consolidated_ambulance_insurance_company_format_xls)
        
        self.consolidated_ambulance_insurance_company = {'kod_smo': '',             # Код страховой компании
                                                         'current_month': '',       # Месяц подготовки сводного счета
                                                         'year_bill': '',           # Отчетный год
                                                         'number_consolidated': '', # Номер сводного, он же номер месяца
                                                         'kod_lpu': [],             # Код ЛПУ оказания
                                                         'profil': [],              # Профиль отделения оказания помощи
                                                         'kod_usl': [],             # Код посещения, услуги
                                                         'visits': [],              # Количество посещений
                                                         'appeal': [],              # Количество обращений
                                                         'services_in_sluch': [],   # Количество услуг в случае
                                                         'dentistry_uet': [],       # Количество УЕТ стоматологии
                                                         'individual_bill': [],     # Количество индивидуальных счетов
                                                         'summ_sluch': []}          # Стоимость услуги
        
        self.consolidated_ambulance_insurance_company['current_month'] = list_consolidated_smo[0]['current_month']
        self.consolidated_ambulance_insurance_company['year_bill'] = list_consolidated_smo[0]['year_bill']
        self.consolidated_ambulance_insurance_company['number_consolidated'] = list_consolidated_smo[0]['number_consolidated']
        
        # Сводный без ТФОМС
        self._consolidated_without_tfoms(settings, list_consolidated_smo)
        self._consolidated_without_tfoms_file(settings)
        # Сводный c ТФОМС
        super()._clear_dict_ambulance()
        self._consolidated_with_tfoms(settings, list_consolidated_smo)
        self._consolidated_with_tfoms_file(settings)
        
        
        return None

    def _consolidated_without_tfoms_file(self, settings):
        out_file = openpyxl.Workbook()
        self.out_sheet = out_file.active
        self.out_sheet.title = 'Сводный'
        super()._bill_formation(self.format_for_smo, settings.medical_organization_data,
                              settings.insurance_company_data,  self.consolidated_ambulance_insurance_company)
        self.start_row = int(self.format_for_smo.getElementsByTagName("start_row")[0].childNodes[0].data)
        self._format_consolidated(self.format_for_smo, self.consolidated_ambulance_insurance_company, settings.medical_organization_data, settings.insurance_company_data, settings.dict_ambulance_profiles)
        if not os.path.exists(settings.source_dir + '/поликлиника'):
            os.mkdir(settings.source_dir + '/поликлиника')
        out_file.save(settings.source_dir + 'поликлиника/' +
                           'сводный_' + 'без ТФОМС' + '.xlsx')
        out_file.close()

    def _consolidated_without_tfoms(self, settings, list_consolidated_smo):
        for consolidated_smo in list_consolidated_smo:
            if consolidated_smo['kod_smo'] == settings.code_tfoms:
                continue
            self._consolidated_dict(settings, consolidated_smo)
    
    def _consolidated_with_tfoms_file(self, settings):
        out_file = openpyxl.Workbook()
        self.out_sheet = out_file.active
        self.out_sheet.title = 'Сводный'
        super()._bill_formation(self.format_for_smo, settings.medical_organization_data,
                              settings.insurance_company_data,  self.consolidated_ambulance_insurance_company)
        self.start_row = int(self.format_for_smo.getElementsByTagName("start_row")[0].childNodes[0].data)
        self._format_consolidated(self.format_for_smo, self.consolidated_ambulance_insurance_company, settings.medical_organization_data, settings.insurance_company_data, settings.dict_ambulance_profiles)
        if not os.path.exists(settings.source_dir + '/поликлиника'):
            os.mkdir(settings.source_dir + '/поликлиника')
        out_file.save(settings.source_dir + 'поликлиника/' +
                           'сводный_' + 'с ТФОМС' + '.xlsx')
        out_file.close()

    def _consolidated_with_tfoms(self, settings, list_consolidated_smo):
        for consolidated_smo in list_consolidated_smo:
            if consolidated_smo['kod_smo'] != settings.code_tfoms:
                continue
            self._consolidated_dict(settings, consolidated_smo)
    
    def _consolidated_dict(self, settings, consolidated_smo):
        consolidated_sluch = {'kod_lpu': None,          # 0 Код ЛПУ оказания
                              'profil': None,           # 1 Профиль отделения оказания помощи
                              'kod_usl': None,          # 2 Код посещения, услуги
                              'visits': 0,              # 3 Количество посещений
                              'appeal': 0,              # 4 Количество обращений
                              'services_in_sluch': 0,   # 5 Количество услуг в случае
                              'dentistry_uet': 0,       # 6 Количество УЕТ стоматологии
                              'individual_bill': 1,     # 7 Количество индивидуальных счетов
                              'summ_sluch': 0}          # 8 Стоимость услуги
        
        for index_kod_lpu in range(len(consolidated_smo['kod_lpu'])):
            for index_profil in range(len(consolidated_smo['profil'][index_kod_lpu])):
                for index_kod_usl in range(len(consolidated_smo['kod_usl'][index_kod_lpu][index_profil])):
                    consolidated_sluch['kod_lpu'] = consolidated_smo['kod_lpu'][index_kod_lpu]
                    consolidated_sluch['profil'] = consolidated_smo['profil'][index_kod_lpu][index_profil]
                    consolidated_sluch['kod_usl'] = consolidated_smo['kod_usl'][index_kod_lpu][index_profil][index_kod_usl]
                    consolidated_sluch['visits'] = consolidated_smo['visits'][index_kod_lpu][index_profil][index_kod_usl]
                    consolidated_sluch['appeal'] = consolidated_smo['appeal'][index_kod_lpu][index_profil][index_kod_usl]
                    consolidated_sluch['services_in_sluch'] = consolidated_smo['services_in_sluch'][index_kod_lpu][index_profil][index_kod_usl]
                    consolidated_sluch['dentistry_uet'] = consolidated_smo['dentistry_uet'][index_kod_lpu][index_profil][index_kod_usl]
                    consolidated_sluch['individual_bill'] = consolidated_smo['individual_bill'][index_kod_lpu][index_profil][index_kod_usl]
                    consolidated_sluch['summ_sluch'] = consolidated_smo['summ_sluch'][index_kod_lpu][index_profil][index_kod_usl]
                    self.__consolidated_ambulance_add_finded_case(consolidated_sluch)
    
    def __consolidated_ambulance_add_finded_case(self, consolidated_sluch):
        """
        Добавляем информацию о найденном случае в общий словарь поликлиники
        """
        address_cell_kod_lpu = -1
        address_cell_profil = -1
        address_cell_kod_usl = -1
        
        # Ищем ЛПУ в словаре
        if consolidated_sluch['kod_lpu'] in self.consolidated_ambulance_insurance_company['kod_lpu']:
            # ЛПУ есть в словаре, ищем позицию
            for cell_kod_lpu in range(len(self.consolidated_ambulance_insurance_company['kod_lpu'])):
                if self.consolidated_ambulance_insurance_company['kod_lpu'][cell_kod_lpu] == consolidated_sluch['kod_lpu']:
                    address_cell_kod_lpu = cell_kod_lpu
                    break
        else:
            # ЛПУ нет в словаре, добавляем и запоминаем позицию
            address_cell_kod_lpu = self.__consolidated_ambulance_insurance_company_add_new_lpu(consolidated_sluch['kod_lpu'], consolidated_sluch['profil'], consolidated_sluch['kod_usl'])
        
        # Ищем профиль в словаре
        if address_cell_profil == -1 and (consolidated_sluch['profil'] in self.consolidated_ambulance_insurance_company['profil'][address_cell_kod_lpu]):
            # профиль есть в списке, ищем позицию
            for cell_profil in range(0, len(self.consolidated_ambulance_insurance_company['profil'][address_cell_kod_lpu])):
                if self.consolidated_ambulance_insurance_company['profil'][address_cell_kod_lpu][cell_profil] == consolidated_sluch['profil']:
                    address_cell_profil = cell_profil
                    break
        else:
            # профиля нет в списке, добавляем
            address_cell_profil = self.__consolidated_ambulance_insurance_company_add_new_profil(consolidated_sluch['profil'], consolidated_sluch['kod_usl'], address_cell_kod_lpu)
        
        # Ищем код услуги в списке
        if address_cell_kod_usl == -1 and consolidated_sluch['kod_usl'] in self.consolidated_ambulance_insurance_company['kod_usl'][address_cell_kod_lpu][address_cell_profil]:
            # код услуги есть в списке, ищем позицию
            for cell_kod_usl in range(len(self.consolidated_ambulance_insurance_company['kod_usl'][address_cell_kod_lpu][address_cell_profil])):
                if self.consolidated_ambulance_insurance_company['kod_usl'][address_cell_kod_lpu][address_cell_profil][cell_kod_usl] == consolidated_sluch['kod_usl']:
                    address_cell_kod_usl = cell_kod_usl
                    break
        else:
            # кода услуги нет в списке, добавляем
            address_cell_kod_usl = self.__consolidated_ambulance_insurance_company_add_new_usl(consolidated_sluch['kod_usl'], address_cell_kod_lpu, address_cell_profil)
        # прибавляем данные для лп, профиля, кода услуги
        self.consolidated_ambulance_insurance_company['visits'][address_cell_kod_lpu][address_cell_profil][address_cell_kod_usl] += consolidated_sluch['visits']
        self.consolidated_ambulance_insurance_company['appeal'][address_cell_kod_lpu][address_cell_profil][address_cell_kod_usl] += consolidated_sluch['appeal']
        self.consolidated_ambulance_insurance_company['services_in_sluch'][address_cell_kod_lpu][address_cell_profil][address_cell_kod_usl] += consolidated_sluch['services_in_sluch']
        self.consolidated_ambulance_insurance_company['dentistry_uet'][address_cell_kod_lpu][address_cell_profil][address_cell_kod_usl] += consolidated_sluch['dentistry_uet']
        self.consolidated_ambulance_insurance_company['individual_bill'][address_cell_kod_lpu][address_cell_profil][address_cell_kod_usl] += consolidated_sluch['individual_bill']
        self.consolidated_ambulance_insurance_company['summ_sluch'][address_cell_kod_lpu][address_cell_profil][address_cell_kod_usl] += consolidated_sluch['summ_sluch']

    def __consolidated_ambulance_insurance_company_add_new_lpu(self, kod_lpu, profil, kod_usl):
        """
        Добавляем новую больницу в общий словарь поликлиники
        """
        self.consolidated_ambulance_insurance_company['kod_lpu'].append(kod_lpu)
        address_cell_kod_lpu = len(self.consolidated_ambulance_insurance_company['kod_lpu']) - 1
        self.consolidated_ambulance_insurance_company['profil'].append([])
        self.consolidated_ambulance_insurance_company['kod_usl'].append([])
        self.consolidated_ambulance_insurance_company['visits'].append([])
        self.consolidated_ambulance_insurance_company['appeal'].append([])
        self.consolidated_ambulance_insurance_company['services_in_sluch'].append([])
        self.consolidated_ambulance_insurance_company['dentistry_uet'].append([])
        self.consolidated_ambulance_insurance_company['individual_bill'].append([])
        self.consolidated_ambulance_insurance_company['summ_sluch'].append([])
        self.__consolidated_ambulance_insurance_company_add_new_profil(profil, kod_usl, address_cell_kod_lpu)
        return address_cell_kod_lpu
    
    def __consolidated_ambulance_insurance_company_add_new_profil(self, profil, kod_usl, address_cell_kod_lpu):
        """
        Добавляем новое подразделение в общий словарь поликлиники
        """
        self.consolidated_ambulance_insurance_company['profil'][address_cell_kod_lpu].append(profil)
        address_cell_profil = len(self.consolidated_ambulance_insurance_company['profil'][address_cell_kod_lpu]) - 1
        self.consolidated_ambulance_insurance_company['kod_usl'][address_cell_kod_lpu].append([])
        self.consolidated_ambulance_insurance_company['visits'][address_cell_kod_lpu].append([])
        self.consolidated_ambulance_insurance_company['appeal'][address_cell_kod_lpu].append([])
        self.consolidated_ambulance_insurance_company['services_in_sluch'][address_cell_kod_lpu].append([])
        self.consolidated_ambulance_insurance_company['dentistry_uet'][address_cell_kod_lpu].append([])
        self.consolidated_ambulance_insurance_company['individual_bill'][address_cell_kod_lpu].append([])
        self.consolidated_ambulance_insurance_company['summ_sluch'][address_cell_kod_lpu].append([])
        self.__consolidated_ambulance_insurance_company_add_new_usl(kod_usl, address_cell_kod_lpu, address_cell_profil)
        return address_cell_profil
    
    def __consolidated_ambulance_insurance_company_add_new_usl(self, kod_usl, address_cell_kod_lpu, address_cell_profil):
        """
        Добавляем новую услугу в подразделение в общий словарь поликлиники
        """
        self.consolidated_ambulance_insurance_company['kod_usl'][address_cell_kod_lpu][address_cell_profil].append(kod_usl)
        self.consolidated_ambulance_insurance_company['visits'][address_cell_kod_lpu][address_cell_profil].append(0)
        self.consolidated_ambulance_insurance_company['appeal'][address_cell_kod_lpu][address_cell_profil].append(0)
        self.consolidated_ambulance_insurance_company['services_in_sluch'][address_cell_kod_lpu][address_cell_profil].append(0)
        self.consolidated_ambulance_insurance_company['dentistry_uet'][address_cell_kod_lpu][address_cell_profil].append(0)
        self.consolidated_ambulance_insurance_company['individual_bill'][address_cell_kod_lpu][address_cell_profil].append(0)
        self.consolidated_ambulance_insurance_company['summ_sluch'][address_cell_kod_lpu][address_cell_profil].append(0)
        return len(self.consolidated_ambulance_insurance_company['kod_usl'][address_cell_kod_lpu][address_cell_profil]) - 1

    def _change_value_if_var(self, value, medical_organization_data, insurance_company_data, consolidated_insurance_company):
        """
        Получает название переменной. возвражает значение переменной и данных счета
        """
        var = value[3:]
        if var.find('_mo_') != -1:
            var = var[4:]
            return medical_organization_data[var]
        if var.find('_smo_') != -1:
            try:
                smo_list_index = insurance_company_data['kod_smo'].index(consolidated_insurance_company['kod_smo'])
                var = var[5:]
                value = insurance_company_data[var][smo_list_index]
            except ValueError:
                value = ''
            return value
        if var.find('_consolidated_') != -1:
            var = var[14:]
            return consolidated_insurance_company[var]
        return value
    
    def _format_consolidated(self, format_for_smo, consolidated_insurance_company, medical_organization_data, insurance_company_data, names_profiles):
        """
        Заполняем данные сводного счета по подразделениям,
        профилям и услугам
        """
        self.current_row = self.start_row
        for index_lpu in range(len(consolidated_insurance_company['kod_lpu'])):
            self.start_row = self.current_row
            for index_profil in range(len(consolidated_insurance_company['profil'][index_lpu])):
                for index_kod_usl in range(len(consolidated_insurance_company['kod_usl'][index_lpu][index_profil])):
                    for cell_format in format_for_smo.getElementsByTagName("rcell"):
                        first_cell = super()._format_cell_consolidated(cell_format)
                        # Запись значения
                        try:
                            value_for_cell = cell_format.getElementsByTagName("value")[0].childNodes[0].data
                        except IndexError:
                            value_for_cell = ''
                        if value_for_cell.find('var_') != -1:
                            value_for_cell = self._change_value_consolidated_if_var(value_for_cell, consolidated_insurance_company, index_lpu, index_profil, index_kod_usl, names_profiles)
                        self.out_sheet[first_cell].value = value_for_cell
                    self.current_row += 1
            # Предварительные результаты
            for cell_format in format_for_smo.getElementsByTagName("pecell"):
                first_cell = super()._format_cell_consolidated(cell_format)
                column = cell_format.getElementsByTagName("first_cell")[0].childNodes[0].data
                # Запись значения
                try:
                    value_for_cell = cell_format.getElementsByTagName("value")[0].childNodes[0].data
                except IndexError:
                    value_for_cell = ''
                if value_for_cell.find('var_') != -1:
                    value_for_cell = super()._change_value_consolidated_ambulance_preresult_if_var(value_for_cell, column, consolidated_insurance_company['kod_lpu'][index_lpu])
                self.out_sheet[first_cell].value = value_for_cell
            self.current_row += 1
        
        # Окончательные результаты
        for cell_format in format_for_smo.getElementsByTagName("ecell"):
            first_cell = super()._format_cell_consolidated(cell_format)
            # Запись значения
            try:
                value_for_cell = cell_format.getElementsByTagName("value")[0].childNodes[0].data
            except IndexError:
                value_for_cell = ''
            if value_for_cell.find('var_') != -1:
                value_for_cell = super()._change_value_consolidated_ambulance_result_if_var(value_for_cell, column)
            self.out_sheet[first_cell].value = value_for_cell
        self.current_row += 3
        # Завершающая строка
        for cell_format in format_for_smo.getElementsByTagName("edcell"):
            first_cell = super()._format_cell_consolidated(cell_format)
            # Запись значения
            try:
                value_for_cell = cell_format.getElementsByTagName("value")[0].childNodes[0].data
            except IndexError:
                value_for_cell = ''
            if value_for_cell.find('summ_final_cell') != -1:
                value_for_cell = '=' + self.cell_final_summ
            self.out_sheet[first_cell].value = value_for_cell


class ConsolidatedHospitalBillGenerator(FormatGenerator):
    """
    Сводыные счета по поликлиникие с ТФОМС и без
    """
    def __init__(self, *args, **kwargs):
        """
        Инициализация отчетов
        """
        try:
            settings = kwargs['settings']
        except KeyError:
            print('Для формирования счета необходимо передать настройки')
            sys.exit(1)
        try:
            list_consolidated_smo = kwargs['list_consolidated_smo']
        except KeyError:
            print('Не получены данные для формирования сводного счета')
            sys.exit(1)
        self.format_for_smo = super()._select_format_for_smo(settings.consolidated_hospital_insurance_company_format_xls)
        
        self.consolidated_insurance_company = {'kod_smo': '',             # Код страховой компании
                                               'number_value': 1,
                                               'current_month': '',       # Месяц подготовки сводного счета
                                               'year_bill': '',           # Отчетный год
                                               'number_consolidated': '', # Номер сводного, он же номер месяца
                                               'kod_lpu': [],             # 0 Код ЛПУ оказания
                                               'podr': [],                # 1 Код подразделения оказания помощи
                                               'pacients': [],            # 3 Колчисевто выбывших пациентов
                                               'amount_of_days': [],      # 4 Количество койкодней фактическое
                                               'amount_of_days_paid': [], # 5 Количество койкодней оплачено
                                               'fksg': [],                # 2 Код примененного КСГ, поле code_usl
                                               'summ': [],                # 6 Сумма за оказанную помощь
                                               'type_hospital': ''}       # 7 Вид стационарной помощий
        
        self.consolidated_insurance_company['current_month'] = list_consolidated_smo[0]['current_month']
        self.consolidated_insurance_company['year_bill'] = list_consolidated_smo[0]['year_bill']
        self.consolidated_insurance_company['number_consolidated'] = list_consolidated_smo[0]['number_consolidated']
        self.consolidated_insurance_company['type_hospital'] = list_consolidated_smo[0]['type_hospital']
        
        if self.consolidated_insurance_company['type_hospital'] == 'круглосуточного стационара':
            self.path_dir = '/стационар_круглосуточный/'
        elif self.consolidated_insurance_company['type_hospital'] == 'дневного стационара':
            self.path_dir = '/стационар_дневной/'
        else:
            self.path_dir = '/стационар/'
        
        # Сводный без ТФОМС
        self._consolidated_without_tfoms(settings, list_consolidated_smo)
        self._consolidated_without_tfoms_file(settings)
        # Сводный c ТФОМС
        super()._clear_dict_hospital()
        super()._clear_dict_hospital_subtotal()
        self._consolidated_with_tfoms(settings, list_consolidated_smo)
        self._consolidated_with_tfoms_file(settings)
        return None

    def _change_value_if_var(self, value, medical_organization_data, insurance_company_data, consolidated_insurance_company):
        """
        Получает название переменной. возвражает значение переменной и данных счета
        """
        var = value[3:]
        if var.find('_mo_') != -1:
            var = var[4:]
            return medical_organization_data[var]
        if var.find('_smo_') != -1:
            try:
                smo_list_index = insurance_company_data['kod_smo'].index(consolidated_insurance_company['kod_smo'])
                var = var[5:]
                value = insurance_company_data[var][smo_list_index]
            except ValueError:
                value = ''
            return value
        if var.find('_consolidated_') != -1:
            var = var[14:]
            return str(consolidated_insurance_company[var])
        return value
    
    def _format_consolidated(self, format_for_smo, consolidated_insurance_company, medical_organization_data, insurance_company_data, names_profiles):
        """
        Заполняем данные сводного счета по подразделениям,
        профилям и услугам
        """
        self.current_row = self.start_row
        for index_lpu in range(len(consolidated_insurance_company['kod_lpu'])):
            for index_podr in range(len(consolidated_insurance_company['podr'][index_lpu])):
                self.start_row = self.current_row
                for index_fksg in range(len(consolidated_insurance_company['fksg'][index_lpu][index_podr])):
                    for cell_format in format_for_smo.getElementsByTagName("rcell"):
                        first_cell = super()._format_cell_consolidated(cell_format)
                        # Запись значения
                        try:
                            value_for_cell = cell_format.getElementsByTagName("value")[0].childNodes[0].data
                        except IndexError:
                            value_for_cell = ''
                        if value_for_cell.find('var_') != -1:
                            value_for_cell = self._change_value_consolidated_if_var(value_for_cell, consolidated_insurance_company, index_lpu, index_podr, index_fksg, names_profiles)
                        self.out_sheet[first_cell].value = value_for_cell
                    consolidated_insurance_company['number_value'] += 1
                    self.current_row += 1
                    
                # Предварительные результаты отделения
                for cell_format in format_for_smo.getElementsByTagName("pecell_podr"):
                    first_cell = super()._format_cell_consolidated(cell_format)
                    column = cell_format.getElementsByTagName("first_cell")[0].childNodes[0].data
                    # Запись значения
                    try:
                        value_for_cell = cell_format.getElementsByTagName("value")[0].childNodes[0].data
                    except IndexError:
                        value_for_cell = ''
                    if value_for_cell.find('var_') != -1:
                        value_for_cell = super()._change_value_consolidated_hospital_preresult_if_var(value_for_cell, column, names_profiles[str(consolidated_insurance_company['podr'][index_lpu][index_podr])[:4]])
                    self.out_sheet[first_cell].value = value_for_cell
                self.current_row += 1
                
            # Предварительные результаты ЛПУ
            for cell_format in format_for_smo.getElementsByTagName("pecell_lpu"):
                first_cell = super()._format_cell_consolidated(cell_format)
                column = cell_format.getElementsByTagName("first_cell")[0].childNodes[0].data
                # Запись значения
                try:
                    value_for_cell = cell_format.getElementsByTagName("value")[0].childNodes[0].data
                except IndexError:
                    value_for_cell = ''
                if value_for_cell.find('var_') != -1:
                    value_for_cell = super()._change_value_consolidated_hosppital_subtotal_if_var(value_for_cell, column, consolidated_insurance_company['kod_lpu'][index_lpu])
                self.out_sheet[first_cell].value = value_for_cell
            self.current_row += 1
            self._clear_dict_hospital()
        
        # Окончательные результаты
        for cell_format in format_for_smo.getElementsByTagName("ecell"):
            first_cell = super()._format_cell_consolidated(cell_format)
            # Запись значения
            try:
                value_for_cell = cell_format.getElementsByTagName("value")[0].childNodes[0].data
            except IndexError:
                value_for_cell = ''
            if value_for_cell.find('var_') != -1:
                value_for_cell = super()._change_value_consolidated_hospital_result_if_var(value_for_cell, column)
            self.out_sheet[first_cell].value = value_for_cell
        self.current_row += 3
        # Завершающая строка
        for cell_format in format_for_smo.getElementsByTagName("edcell"):
            first_cell = super()._format_cell_consolidated(cell_format)
            # Запись значения
            try:
                value_for_cell = cell_format.getElementsByTagName("value")[0].childNodes[0].data
            except IndexError:
                value_for_cell = ''
            if value_for_cell.find('summ_final_cell') != -1:
                value_for_cell = '=' + self.cell_final_summ
            self.out_sheet[first_cell].value = value_for_cell

    def _consolidated_without_tfoms(self, settings, list_consolidated_smo):
        for consolidated_smo in list_consolidated_smo:
            if consolidated_smo['kod_smo'] == settings.code_tfoms:
                continue
            self._consolidated_dict(settings, consolidated_smo)
    
    def _consolidated_without_tfoms_file(self, settings):
        out_file = openpyxl.Workbook()
        self.out_sheet = out_file.active
        self.out_sheet.title = 'Сводный'
        super()._bill_formation(self.format_for_smo, settings.medical_organization_data,
                              settings.insurance_company_data,  self.consolidated_insurance_company)
        self.start_row = int(self.format_for_smo.getElementsByTagName("start_row")[0].childNodes[0].data)
        self._format_consolidated(self.format_for_smo, self.consolidated_insurance_company, settings.medical_organization_data, settings.insurance_company_data, settings.dict_hospital_profiles)
        if not os.path.exists(settings.source_dir + self.path_dir):
            os.mkdir(settings.source_dir + self.path_dir)
        out_file.save(settings.source_dir + self.path_dir +
                           'сводный_' + self.consolidated_insurance_company['type_hospital'] + 'без ТФОМС' + '.xlsx')
        out_file.close()
    
    def _consolidated_with_tfoms(self, settings, list_consolidated_smo):
        for consolidated_smo in list_consolidated_smo:
            if consolidated_smo['kod_smo'] != settings.code_tfoms:
                continue
            self._consolidated_dict(settings, consolidated_smo)
    
    def _consolidated_with_tfoms_file(self, settings):
        out_file = openpyxl.Workbook()
        self.out_sheet = out_file.active
        self.out_sheet.title = 'Сводный'
        super()._bill_formation(self.format_for_smo, settings.medical_organization_data,
                              settings.insurance_company_data,  self.consolidated_insurance_company)
        self.start_row = int(self.format_for_smo.getElementsByTagName("start_row")[0].childNodes[0].data)
        self._format_consolidated(self.format_for_smo, self.consolidated_insurance_company, settings.medical_organization_data, settings.insurance_company_data, settings.dict_hospital_profiles)
        if not os.path.exists(settings.source_dir + self.path_dir):
            os.mkdir(settings.source_dir + self.path_dir)
        out_file.save(settings.source_dir + self.path_dir +
                           'сводный_' + self.consolidated_insurance_company['type_hospital'] + 'с ТФОМС' + '.xlsx')
        out_file.close()
    
    def _consolidated_dict(self, settings, consolidated_smo):
        consolidated_sluch = {'kod_lpu': None,             # 0 Код ЛПУ оказания
                              'podr': None,                # 1 Код подразделения оказания помощи
                              'pacients': 0,               # 3 Колчисевто выбывших пациентов
                              'amount_of_days': None,      # 4 Количество койкодней фактическое
                              'amount_of_days_paid': None, # 5 Количество койкодней оплачено
                              'fksg': None,                # 2 Код примененного КСГ, поле code_usl
                              'summ_usl': 0}               # 6 Сумма за Услугу
        
        for index_kod_lpu in range(len(consolidated_smo['kod_lpu'])):
            for index_podr in range(len(consolidated_smo['podr'][index_kod_lpu])):
                for index_fksg in range(len(consolidated_smo['fksg'][index_kod_lpu][index_podr])):
                    consolidated_sluch['kod_lpu'] = consolidated_smo['kod_lpu'][index_kod_lpu]
                    consolidated_sluch['podr'] = consolidated_smo['podr'][index_kod_lpu][index_podr]
                    consolidated_sluch['fksg'] = consolidated_smo['fksg'][index_kod_lpu][index_podr][index_fksg]
                    consolidated_sluch['pacients'] = consolidated_smo['pacients'][index_kod_lpu][index_podr][index_fksg]
                    consolidated_sluch['amount_of_days'] = consolidated_smo['amount_of_days'][index_kod_lpu][index_podr][index_fksg]
                    consolidated_sluch['amount_of_days_paid'] = consolidated_smo['amount_of_days_paid'][index_kod_lpu][index_podr][index_fksg]
                    consolidated_sluch['summ_usl'] = consolidated_smo['summ'][index_kod_lpu][index_podr][index_fksg]
                    self.__consolidated_add_finded_case(consolidated_sluch)
    
    def __consolidated_add_finded_case(self, consolidated_insurance_company):
        """
        Добавляем информацию о найденном случае в общий словарь дневного стационара
        """
        address_cell_kod_lpu = -1
        address_cell_podr = -1
        address_cell_fksg = -1
        
        # Ищем ЛПУ в словаре
        if consolidated_insurance_company['kod_lpu'] in self.consolidated_insurance_company['kod_lpu']:
            # ЛПУ есть в словаре, ищем позицию
            for cell_kod_lpu in range(len(self.consolidated_insurance_company['kod_lpu'])):
                if self.consolidated_insurance_company['kod_lpu'][cell_kod_lpu] == consolidated_insurance_company['kod_lpu']:
                    address_cell_kod_lpu = cell_kod_lpu
                    break
        else:
            # ЛПУ нет в словаре, добавляем и запоминаем позицию
            address_cell_kod_lpu = self.__consolidated_insurance_company_add_new_lpu(consolidated_insurance_company['kod_lpu'], consolidated_insurance_company['podr'], consolidated_insurance_company['fksg'])
        
        # Ищем профиль в словаре
        if address_cell_podr == -1 and (consolidated_insurance_company['podr'] in self.consolidated_insurance_company['podr'][address_cell_kod_lpu]):
            # профиль есть в списке, ищем позицию
            for cell_podr in range(0, len(self.consolidated_insurance_company['podr'][address_cell_kod_lpu])):
                if self.consolidated_insurance_company['podr'][address_cell_kod_lpu][cell_podr] == consolidated_insurance_company['podr']:
                    address_cell_podr = cell_podr
                    break
        else:
            # профиля нет в списке, добавляем
            address_cell_podr = self.__consolidated_insurance_company_add_new_podr(consolidated_insurance_company['podr'], consolidated_insurance_company['fksg'], address_cell_kod_lpu)
        
        # Ищем код услуги в списке
        if address_cell_fksg == -1 and consolidated_insurance_company['fksg'] in self.consolidated_insurance_company['fksg'][address_cell_kod_lpu][address_cell_podr]:
            # код услуги есть в списке, ищем позицию
            for cell_fksg in range(len(self.consolidated_insurance_company['fksg'][address_cell_kod_lpu][address_cell_podr])):
                if self.consolidated_insurance_company['fksg'][address_cell_kod_lpu][address_cell_podr][cell_fksg] == consolidated_insurance_company['fksg']:
                    address_cell_fksg = cell_fksg
                    break
        else:
            # кода услуги нет в списке, добавляем
            address_cell_fksg = self.__consolidated_insurance_company_add_new_fksg(consolidated_insurance_company['fksg'], address_cell_kod_lpu, address_cell_podr)
        # прибавляем данные для лп, профиля, кода услуги
        self.consolidated_insurance_company['pacients'][address_cell_kod_lpu][address_cell_podr][address_cell_fksg] += consolidated_insurance_company['pacients']
        self.consolidated_insurance_company['amount_of_days'][address_cell_kod_lpu][address_cell_podr][address_cell_fksg] += consolidated_insurance_company['amount_of_days']
        self.consolidated_insurance_company['amount_of_days_paid'][address_cell_kod_lpu][address_cell_podr][address_cell_fksg] += consolidated_insurance_company['amount_of_days_paid']
        self.consolidated_insurance_company['summ'][address_cell_kod_lpu][address_cell_podr][address_cell_fksg] += consolidated_insurance_company['summ_usl']
    
    def __consolidated_insurance_company_add_new_lpu(self, kod_lpu, podr, fksg):
        """
        Добавляем новую больницу в общий словарь стационара
        """
        self.consolidated_insurance_company['kod_lpu'].append(kod_lpu)
        address_cell_kod_lpu = len(self.consolidated_insurance_company['kod_lpu']) - 1
        self.consolidated_insurance_company['podr'].append([])
        self.consolidated_insurance_company['fksg'].append([])
        self.consolidated_insurance_company['pacients'].append([])
        self.consolidated_insurance_company['amount_of_days'].append([])
        self.consolidated_insurance_company['amount_of_days_paid'].append([])
        self.consolidated_insurance_company['summ'].append([])
        self.__consolidated_insurance_company_add_new_podr(podr, fksg, address_cell_kod_lpu)
        return address_cell_kod_lpu
    
    def __consolidated_insurance_company_add_new_podr(self, podr, fksg, address_cell_kod_lpu):
        """
        Добавляем новое подразделение в общий словарь стационара
        """
        self.consolidated_insurance_company['podr'][address_cell_kod_lpu].append(podr)
        address_cell_podr = len(self.consolidated_insurance_company['podr'][address_cell_kod_lpu]) - 1
        self.consolidated_insurance_company['fksg'][address_cell_kod_lpu].append([])
        self.consolidated_insurance_company['pacients'][address_cell_kod_lpu].append([])
        self.consolidated_insurance_company['amount_of_days'][address_cell_kod_lpu].append([])
        self.consolidated_insurance_company['amount_of_days_paid'][address_cell_kod_lpu].append([])
        self.consolidated_insurance_company['summ'][address_cell_kod_lpu].append([])
        self.__consolidated_insurance_company_add_new_fksg(fksg, address_cell_kod_lpu, address_cell_podr)
        return address_cell_podr
    
    def __consolidated_insurance_company_add_new_fksg(self, fksg, address_cell_kod_lpu, address_cell_podr):
        """
        Добавляем новую услугу в подразделение в общий словарь стационара
        """
        self.consolidated_insurance_company['fksg'][address_cell_kod_lpu][address_cell_podr].append(fksg)
        self.consolidated_insurance_company['pacients'][address_cell_kod_lpu][address_cell_podr].append(0)
        self.consolidated_insurance_company['amount_of_days'][address_cell_kod_lpu][address_cell_podr].append(0)
        self.consolidated_insurance_company['amount_of_days_paid'][address_cell_kod_lpu][address_cell_podr].append(0)
        self.consolidated_insurance_company['summ'][address_cell_kod_lpu][address_cell_podr].append(0)
        return len(self.consolidated_insurance_company['fksg'][address_cell_kod_lpu][address_cell_podr]) - 1


class ConsolidatedAMbulanceBillBaseGenerator(FormatGenerator):
    """
    Формирование сводного счета по поликлинике в разрезе основных данных,
    страховая, количество обращений, количество посещений, УЕТ, сумма
    """
    def __init__(self):
        pass
